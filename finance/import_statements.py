#!/usr/bin/env python3
"""
Import July–August 2026 bank / CC statements into the live finance workbook Ledger.

Sources (under finance/):
  - hdfc Acct_Statement_*7754_06082026.xls        → HDFC Savings Jul
  - hdfc Acct_Statement_*7754_06082026(1).xls     → HDFC Savings Aug 1–6
  - icici OpTransactionHistory06-08-2026.xls      → ICICI Savings Jul
  - icici CCStatement_Current06-08-2026.xls       → ICICI Credit Card
  - Hard-coded HDFC Credit Card rows from phone UI

Opening balances are set so calculated balances match statement closings.
"""
from __future__ import annotations

import csv
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
# Live workbook under Documents (not the repo copy).
WORKBOOK = Path("/home/himanshu/Documents/Finance/Finance-Mng-V2.xlsx")
BACKUP = WORKBOOK.with_name("Finance-Mng-V2.pre-import-backup.xlsx")

yellow_fill = PatternFill("solid", fgColor="FFF2CC")
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
inr = "₹#,##0.00"


@dataclass
class Entry:
    date: date
    typ: str
    amount: float
    from_acct: str
    to_acct: str
    category: str
    budget: bool
    recurring: bool = False
    notes: str = ""
    time: time | None = None
    source: str = ""
    # for dual-leg transfer matching (avoid double-count when both banks listed)
    skip: bool = False


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").upper()).strip()


def convert_xls_to_csv(xls_path: Path, out_dir: Path) -> Path:
    """LibreOffice convert (handles classic .xls and OOXML misnamed as .xls)."""
    out_dir.mkdir(parents=True, exist_ok=True)
    # Copy to a clean name so LO output is predictable
    work = out_dir / xls_path.name
    if not work.exists():
        shutil.copy2(xls_path, work)
    result = subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to",
            "csv:Text - txt - csv (StarCalc):44,34,76",
            "--outdir",
            str(out_dir),
            str(work),
        ],
        capture_output=True,
        text=True,
    )
    csv_path = out_dir / (work.stem + ".csv")
    if result.returncode != 0 or not csv_path.exists():
        # Fallback: OOXML via openpyxl (needs real .xlsx suffix)
        head = xls_path.read_bytes()[:4]
        if head == b"PK\x03\x04":
            from openpyxl import load_workbook as _lw

            xlsx_copy = out_dir / (xls_path.stem + ".xlsx")
            shutil.copy2(xls_path, xlsx_copy)
            wb = _lw(xlsx_copy, data_only=True)
            ws = wb.active
            with csv_path.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow(["" if c is None else c for c in row])
            return csv_path
        raise RuntimeError(
            f"convert failed for {xls_path}: {result.stderr or result.stdout}"
        )
    return csv_path


# ── Parsers ──────────────────────────────────────────────────────────────


def parse_hdfc_csv(csv_path: Path) -> list[dict]:
    rows: list[dict] = []
    with csv_path.open(newline="", encoding="utf-8", errors="replace") as f:
        for r in csv.reader(f):
            if not r or not re.match(r"\d{2}/\d{2}/\d{2}", (r[0] or "").strip()):
                continue
            d = datetime.strptime(r[0].strip(), "%d/%m/%y").date()
            narr = r[1].strip() if len(r) > 1 else ""
            w = float(r[4]) if len(r) > 4 and str(r[4]).strip() else 0.0
            dep = float(r[5]) if len(r) > 5 and str(r[5]).strip() else 0.0
            bal = float(r[6]) if len(r) > 6 and str(r[6]).strip() else None
            rows.append(
                {
                    "date": d,
                    "narr": narr,
                    "withdrawal": w,
                    "deposit": dep,
                    "balance": bal,
                    "ref": r[2].strip() if len(r) > 2 else "",
                }
            )
    return rows


def parse_icici_savings_csv(csv_path: Path) -> list[dict]:
    rows: list[dict] = []
    with csv_path.open(newline="", encoding="utf-8", errors="replace") as f:
        for r in csv.reader(f):
            # columns offset by empty first col: S No, Value Date, Txn Date, Cheque, Remarks, Withdrawal, Deposit, Balance
            if len(r) < 8:
                continue
            # find a row where col1 is a serial number
            sno = (r[1] or "").strip() if len(r) > 1 else ""
            if not sno.isdigit():
                continue
            try:
                d = datetime.strptime(str(r[3]).strip(), "%d/%m/%Y").date()
            except Exception:
                continue
            remarks = str(r[5]).strip()
            w = float(str(r[6]).replace(",", "") or 0)
            dep = float(str(r[7]).replace(",", "") or 0)
            bal = float(str(r[8]).replace(",", "") or 0) if len(r) > 8 and str(r[8]).strip() else None
            rows.append(
                {
                    "date": d,
                    "narr": remarks,
                    "withdrawal": w,
                    "deposit": dep,
                    "balance": bal,
                }
            )
    return rows


def parse_icici_cc_csv(csv_path: Path) -> list[dict]:
    """Parse ICICI current statement CSV (openpyxl export or LO)."""
    rows: list[dict] = []
    with csv_path.open(newline="", encoding="utf-8", errors="replace") as f:
        for r in csv.reader(f):
            # find date in any early column
            date_s = None
            details = None
            amount_s = None
            for i, cell in enumerate(r):
                c = str(cell or "").strip()
                if re.match(r"\d{2}-\d{2}-\d{4}", c):
                    date_s = c
                    # details often at i+4 or nearby non-empty
                    for j in range(i + 1, min(i + 8, len(r))):
                        if str(r[j] or "").strip() and not re.match(
                            r"[\d,.]+ (Dr|Cr)\.?", str(r[j]).strip()
                        ):
                            details = str(r[j]).strip()
                            break
                    for j in range(i + 1, min(i + 12, len(r))):
                        m = re.match(
                            r"([\d,.]+)\s*(Dr|Cr)\.?", str(r[j] or "").strip(), re.I
                        )
                        if m:
                            amount_s = m.group(0)
                            break
                    break
            if not date_s or not amount_s:
                continue
            d = datetime.strptime(date_s, "%d-%m-%Y").date()
            m = re.match(r"([\d,.]+)\s*(Dr|Cr)\.?", amount_s, re.I)
            if not m:
                continue
            amt = float(m.group(1).replace(",", ""))
            side = m.group(2).lower()
            rows.append(
                {
                    "date": d,
                    "details": details or "",
                    "amount": amt,
                    "debit": side == "dr",
                    "credit": side == "cr",
                }
            )
    return rows


# ── Categorization ───────────────────────────────────────────────────────


def categorize_hdfc_savings(t: dict) -> Entry:
    d = t["date"]
    n = t["narr"]
    nu = _norm(n)
    w, dep = t["withdrawal"], t["deposit"]
    amt = w or dep
    notes_short = n[:80]

    # Income — salary
    # Bank often credits "July salary" on 31 Jul; user books it as next-month (Aug) income.
    if dep and ("SALARY" in nu or "SAL-EY" in nu or "ACH C- SAL" in nu):
        book = d
        note_extra = ""
        if d.month == 7 and d.year == 2026 and d.day >= 28:
            book = date(2026, 8, 1)
            note_extra = f" | bank date {d.isoformat()} (salary booked as Aug income)"
        return Entry(
            book, "Income", amt, "Employer", "HDFC Savings", "Salary", False,
            notes=f"Salary | {notes_short}{note_extra}", source="hdfc-sav",
        )

    # Credits / inflows
    if dep:
        # INDSTOCKS refund/credit
        if "INDSTOCKS" in nu or "BSE-" in nu:
            return Entry(
                d, "Income", amt, "External", "HDFC Savings", "Cashback", False,
                notes=f"IndStocks/BSE credit | {notes_short}", source="hdfc-sav",
            )
        # Pawan = reimbursement (not salary/income for budget)
        if "PAWAN KUMAR" in nu:
            return Entry(
                d, "Refund", amt, "Expense", "HDFC Savings", "Refund", True,
                notes=f"Reimbursement from Pawan Kumar Gupta | {notes_short}",
                source="hdfc-sav",
            )
        # Vidhisha = loan related (not income)
        if "VIDHISHA" in nu:
            return Entry(
                d, "Transfer", amt, "External", "HDFC Savings", "Transfer", False,
                notes=f"Loan (Vidhisha) — not income | {notes_short}",
                source="hdfc-sav",
            )
        return Entry(
            d, "Income", amt, "External", "HDFC Savings", "Other", False,
            notes=f"Deposit | {notes_short}", source="hdfc-sav",
        )

    # Withdrawals
    # Rent — confirmed ₹10,800 to Santosh Kumar Gupta
    if "SANTOSH KUMAR GUPTA" in nu or (abs(amt - 10800) < 0.01 and "RENT" in nu):
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Rent", False,
            recurring=True, notes=f"Rent (Santosh) | {notes_short}", source="hdfc-sav",
        )

    # ICICI credit card bill pay from HDFC (~11k confirmed)
    if "ICICICREDIT" in nu or "ICICI CREDIT" in nu or ("BILLPAY" in nu and "ICICI" in nu):
        return Entry(
            d, "Credit Card Payment", amt, "HDFC Savings", "ICICI Credit Card",
            "Credit Card Bill", False,
            notes=f"ICICI CC bill pay | {notes_short}", source="hdfc-sav",
        )

    # Self transfer to ICICI (matches ICICI deposit)
    if "HIMANSHU GUPTA" in nu and "ICIC" in nu and amt >= 1000:
        return Entry(
            d, "Transfer", amt, "HDFC Savings", "ICICI Savings", "Transfer", False,
            notes=f"Self transfer to ICICI | {notes_short}", source="hdfc-sav",
        )

    # Jupiter collect — treat as CC/bill payment stack (likely HDFC CC via Jupiter)
    if "JUPITER UPI COLLECT" in nu or "JUPITERUPICOLLECT" in nu:
        return Entry(
            d, "Credit Card Payment", amt, "HDFC Savings", "HDFC Credit Card",
            "Credit Card Bill", False,
            notes=f"Jupiter collect (likely HDFC CC) | {notes_short}", source="hdfc-sav",
        )

    # Amica Financial — EMI-like
    if "AMICA" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "EMIs", False,
            notes=f"Amica Financial | {notes_short}", source="hdfc-sav",
        )

    # Large person-to-person (family)
    if "HENCY PATEL" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Other", False,
            notes=f"To Hency Patel | {notes_short}", source="hdfc-sav",
        )
    if "SONIYA GUPTA" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Other", False,
            notes=f"To Soniya Gupta | {notes_short}", source="hdfc-sav",
        )
    if "XXXXXXX1352" in nu or re.search(r"SBIN0010588", nu):
        # recurring large transfer
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Other", False,
            notes=f"To SBI ****1352 | {notes_short}", source="hdfc-sav",
        )

    # Merchants
    if "VALUE MART" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Groceries - Physical", True,
            notes=f"Value Mart | {notes_short}", source="hdfc-sav",
        )
    if "BLINKIT" in nu or "ZEPTO" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Groceries - Online", True,
            notes=f"Quick commerce | {notes_short}", source="hdfc-sav",
        )
    if "AKSHAYAKALPA" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Groceries - Physical", True,
            recurring=True, notes=f"Akshayakalpa milk | {notes_short}", source="hdfc-sav",
        )
    if "FARMBOWL" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Groceries - Physical", True,
            notes=f"Farmbowl | {notes_short}", source="hdfc-sav",
        )
    if "PROTEIN HUB" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Eating outside", True,
            notes=f"Protein Hub | {notes_short}", source="hdfc-sav",
        )
    if any(x in nu for x in ("SWIGGY", "ZOMATO", "GOKHANA", "TOBOX", "CALIFORNIA BURRITO",
                              "CHOLE BHATU", "KITCHEN BELL", "TELLA CAFE", "REKHA BAR",
                              "GULSHAN E MASTANA", "ANNADORE")):
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Eating outside", True,
            notes=f"Food | {notes_short}", source="hdfc-sav",
        )
    if "URBANCLAP" in nu or "URBAN COMPANY" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "House cleaning (UC)", True,
            notes=f"Urban Company | {notes_short}", source="hdfc-sav",
        )
    if any(x in nu for x in ("BANGALORE METRO", "BMTC", "L AND T METRO", "UBER",
                              "TELANGANA STATE ROAD", "TSRTC", "METRO RAIL")):
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Petrol", True,
            notes=f"Transport | {notes_short}", source="hdfc-sav",
        )
    if "AIRTEL" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Subscription", True,
            notes=f"Airtel | {notes_short}", source="hdfc-sav",
        )
    if "GOOGLE INDIA DIGITAL" in nu or "GPAY-UTILITY" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Subscription", True,
            notes=f"Google / Play utility | {notes_short}", source="hdfc-sav",
        )
    if "VALVE CORPORATION" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Subscription", True,
            notes=f"Steam / Valve | {notes_short}", source="hdfc-sav",
        )
    if "PHARMACY" in nu or "MEDIC" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Medical", False,
            notes=f"Pharmacy/Medical | {notes_short}", source="hdfc-sav",
        )
    if "COMMISSIONER OF POLI" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Other", False,
            notes=f"Police / fine | {notes_short}", source="hdfc-sav",
        )
    if "SWATHI ASSOCIATES" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Other", True,
            notes=f"MS Swathi Associates | {notes_short}", source="hdfc-sav",
        )
    if "M SHIVA KUMAR" in nu:
        # local shop — dosa/chaas style stalls often under this Paytm
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Eating outside", True,
            notes=f"M Shiva Kumar (local) | {notes_short}", source="hdfc-sav",
        )
    if "RAMANI" in nu:
        return Entry(
            d, "Expense", amt, "HDFC Savings", "Expense", "Eating outside", True,
            notes=f"Ramani / local food | {notes_short}", source="hdfc-sav",
        )

    # default small UPI merchant → Other budget
    return Entry(
        d, "Expense", amt, "HDFC Savings", "Expense", "Other", True,
        notes=f"UPI | {notes_short}", source="hdfc-sav",
    )


def categorize_icici_savings(t: dict) -> Entry | None:
    """ICICI savings — skip the inbound self-transfer (already on HDFC side)."""
    d = t["date"]
    n = t["narr"]
    nu = _norm(n)
    w, dep = t["withdrawal"], t["deposit"]
    amt = w or dep

    # Inbound from HDFC self — already recorded as Transfer on HDFC side
    if dep and ("HIMANSHU" in nu or "HDFC" in nu) and amt >= 1000:
        return None  # skip duplicate leg

    # CC bill pay from ICICI savings
    if w and ("CC BILL" in nu or "BILLPAY" in nu or "CC BILLPAY" in nu or re.search(r"CC\s*BILL", nu)):
        return Entry(
            d, "Credit Card Payment", amt, "ICICI Savings", "ICICI Credit Card",
            "Credit Card Bill", False,
            notes=f"ICICI CC bill from ICICI sav | {n[:80]}", source="icici-sav",
        )
    if w and "CC" in nu and "BILL" in nu:
        return Entry(
            d, "Credit Card Payment", amt, "ICICI Savings", "ICICI Credit Card",
            "Credit Card Bill", False,
            notes=f"ICICI CC bill | {n[:80]}", source="icici-sav",
        )

    # Google Play
    if w and ("GOOGLE" in nu or "PLAYSTORE" in nu or "PLAY STORE" in nu):
        return Entry(
            d, "Expense", amt, "ICICI Savings", "Expense", "Subscription", True,
            notes=f"Google Play | {n[:80]}", source="icici-sav",
        )

    if dep:
        return Entry(
            d, "Income", amt, "External", "ICICI Savings", "Other", False,
            notes=f"ICICI deposit | {n[:80]}", source="icici-sav",
        )
    return Entry(
        d, "Expense", amt, "ICICI Savings", "Expense", "Other", True,
        notes=f"ICICI UPI | {n[:80]}", source="icici-sav",
    )


def categorize_icici_cc(t: dict) -> Entry | None:
    d = t["date"]
    det = t["details"]
    du = _norm(det)
    amt = t["amount"]

    # Payment received — already recorded from bank side (HDFC billpay or ICICI sav)
    if t["credit"] and ("PAYMENT" in du or "BBPS" in du):
        return None
    # Refund credit
    if t["credit"]:
        return Entry(
            d, "Refund", amt, "Expense", "ICICI Credit Card", "Refund", True,
            notes=f"ICICI CC refund | {det[:80]}", source="icici-cc",
        )

    # Debits
    if "SWIGGY" in du or "ZOMATO" in du:
        return Entry(
            d, "Expense", amt, "ICICI Credit Card", "Expense", "Eating outside", True,
            notes=f"ICICI CC food | {det[:80]}", source="icici-cc",
        )
    if "BOOKMYSHOW" in du:
        return Entry(
            d, "Expense", amt, "ICICI Credit Card", "Expense", "Other", True,
            notes=f"BookMyShow | {det[:80]}", source="icici-cc",
        )
    if "UTTARAKHAND" in du or "POWER" in du:
        return Entry(
            d, "Expense", amt, "ICICI Credit Card", "Expense", "Electricity bill", True,
            notes=f"Electricity | {det[:80]}", source="icici-cc",
        )
    return Entry(
        d, "Expense", amt, "ICICI Credit Card", "Expense", "Other", True,
        notes=f"ICICI CC | {det[:80]}", source="icici-cc",
    )


# HDFC CC from user phone dump (Jul 29 – Aug 6 2026)
HDFC_CC_MANUAL = [
    (date(2026, 7, 29), 120.00, "Value Mart Super Market Hyderabad", "Groceries - Physical"),
    (date(2026, 7, 30), 303.54, "Sree Balaji Laxmi Naras Medchal", "Medical"),
    (date(2026, 8, 1), 84.00, "Urbanclap Technologies Gurgaon", "House cleaning (UC)"),
    (date(2026, 8, 1), 29871.00, "Interglobviation Gurgaon (IndiGo flight)", "Travel / Flight"),
    (date(2026, 8, 1), 285.33, "Zomato Ltd Bangalore", "Eating outside"),
    (date(2026, 8, 1), 278.00, "Zepto Marketplace Bangalore", "Groceries - Online"),
    (date(2026, 8, 2), 791.00, "Blinkit Bangalore", "Groceries - Online"),
    (date(2026, 8, 3), 170.00, "Tobox Ventures Pvt Ltd Bangalore", "Eating outside"),
    (date(2026, 8, 6), 186.00, "Urbanclap Technologies", "House cleaning (UC)"),
]


def build_entries(tmpdir: Path) -> list[Entry]:
    entries: list[Entry] = []

    # locate source files
    hdfc_jul = next(ROOT.glob("hdfc Acct_Statement_*06082026.xls"))
    hdfc_aug = next(ROOT.glob("hdfc Acct_Statement_*06082026(1).xls"))
    icici_sav = ROOT / "icici OpTransactionHistory06-08-2026.xls"
    icici_cc = ROOT / "icici CCStatement_Current06-08-2026.xls"

    for label, path, parser, mapper in [
        ("hdfc-jul", hdfc_jul, parse_hdfc_csv, categorize_hdfc_savings),
        ("hdfc-aug", hdfc_aug, parse_hdfc_csv, categorize_hdfc_savings),
    ]:
        csv_p = convert_xls_to_csv(path, tmpdir / label)
        for t in parser(csv_p):
            e = mapper(t)
            if e and not e.skip:
                entries.append(e)

    csv_icici_sav = convert_xls_to_csv(icici_sav, tmpdir / "icici-sav")
    for t in parse_icici_savings_csv(csv_icici_sav):
        e = categorize_icici_savings(t)
        if e:
            entries.append(e)

    csv_icici_cc = convert_xls_to_csv(icici_cc, tmpdir / "icici-cc")
    for t in parse_icici_cc_csv(csv_icici_cc):
        e = categorize_icici_cc(t)
        if e:
            entries.append(e)

    for d, amt, note, cat in HDFC_CC_MANUAL:
        budget = cat not in ("Travel / Flight", "Medical", "Rent", "EMIs", "Salary")
        # Medical budget=False in config; Travel/Flight False
        budget_map = {
            "Travel / Flight": False,
            "Medical": False,
            "House cleaning (UC)": True,
            "Eating outside": True,
            "Groceries - Online": True,
            "Groceries - Physical": True,
        }
        entries.append(
            Entry(
                d, "Expense", amt, "HDFC Credit Card", "Expense", cat,
                budget_map.get(cat, True),
                notes=f"HDFC CC | {note}", source="hdfc-cc-manual",
            )
        )

    # sort by date, then source priority (savings before cards on same day), then amount
    source_order = {"hdfc-sav": 0, "icici-sav": 1, "hdfc-cc-manual": 2, "icici-cc": 3}
    entries.sort(key=lambda e: (e.date, source_order.get(e.source, 9), e.amount, e.notes))
    return entries


# ── Opening balances (before first ledger row = 2026-07-01 morning) ──────
# HDFC: Jul 1 first txn DR 30 → bal 88919.52 ⇒ open 88949.52
# ICICI: Jul 1 CR 24000 → bal 24210.77 ⇒ open 210.77
# ICICI CC: statement previous bal 11,723.51 is AFTER the Jul-1 pay of 24,009.66.
#   So open (pre-Jul-1) = 11,723.51 + 24,009.66 = 35,733.17
#   → after Jul pay: 11,723.51; +cycle spends − Aug pay 11,723.51 − refund 140 ≈ 3,795 due.
# HDFC CC: Jupiter Jul payments 23,500.33 + 3,921.48 = 27,421.81 treated as paying prior due.
#   open = 27,421.81 so post-Jupiter due ≈ 0, then + known cycle spends ≈ current due estimate.
OPENING = {
    "HDFC Savings": 88949.52,
    "ICICI Savings": 210.77,
    "Cash": 0.0,
    "Wallet": 0.0,
    "HDFC Credit Card": 27421.81,  # ≈ Jupiter Jul bills paid down
    "ICICI Credit Card": 35733.17,  # 11723.51 prior + 24009.66 Jul-1 payment
    "FD": 0.0,
    "Mutual Fund": 0.0,
    "Employer": 0.0,
    "Expense": 0.0,
    "External": 0.0,
}

# Statement closing targets for verification
TARGETS = {
    "HDFC Savings": 47649.45,  # after last Aug 6 txn
    "ICICI Savings": 142.11,   # after Jul 11 Google Play
    "ICICI Credit Card": 3795.04,  # prev + purchases - pay - refund credit handled
}


def calc_balance(name: str, entries: list[Entry], liability: bool = False) -> float:
    open_bal = OPENING.get(name, 0.0)
    incoming = sum(e.amount for e in entries if e.to_acct == name)
    outgoing = sum(e.amount for e in entries if e.from_acct == name)
    if liability:
        # due = open + charges(from) - payments(to)
        return open_bal + outgoing - incoming
    return open_bal + incoming - outgoing


def write_workbook(entries: list[Entry]) -> dict:
    if not BACKUP.exists():
        shutil.copy2(WORKBOOK, BACKUP)

    wb = load_workbook(WORKBOOK)
    cfg = wb["Configuration"]
    # Opening balances: rows 11–21 map to accounts list order
    account_rows = {
        "HDFC Savings": 11,
        "ICICI Savings": 12,
        "Cash": 13,
        "Wallet": 14,
        "HDFC Credit Card": 15,
        "ICICI Credit Card": 16,
        "FD": 17,
        "Mutual Fund": 18,
        "Employer": 19,
        "Expense": 20,
        "External": 21,
    }
    for name, row in account_rows.items():
        cfg.cell(row, 3, OPENING.get(name, 0.0))
        cfg.cell(row, 3).number_format = inr
        cfg.cell(row, 3).fill = yellow_fill

    cfg["C11"].value = OPENING["HDFC Savings"]
    cfg["F11"] = "Primary savings; open = pre-Jul-1 2026 stmt"
    cfg["F15"] = "Opening due ≈ Jul Jupiter CC bills (paid 27,421.81); refine via recon"
    cfg["F16"] = "Open = 35,733.17 (prior due incl. Jul-1 bill 24,009.66); Aug bill 11,723.51 from HDFC"

    led = wb["Ledger"]
    # Clear existing data rows (keep header)
    if led.max_row > 1:
        led.delete_rows(2, led.max_row - 1)

    # Need enough formula/data rows
    n = len(entries)
    capacity = max(n + 50, 250)

    for i, e in enumerate(entries):
        r = 2 + i
        led.cell(r, 1, e.date).number_format = "dd/mm/yyyy"
        led.cell(r, 1).fill = yellow_fill
        if e.time:
            led.cell(r, 2, e.time).number_format = "HH:mm"
        else:
            led.cell(r, 2, None)
        led.cell(r, 2).fill = yellow_fill
        led.cell(r, 2).number_format = "HH:mm"
        led.cell(r, 3, f'=IF(A{r}="","",TEXT(A{r},"dddd"))')
        led.cell(r, 4, f'=IF(A{r}="","",TEXT(A{r},"MMMM"))')
        led.cell(r, 5, f'=IF(A{r}="","",YEAR(A{r}))')
        led.cell(r, 6, e.typ).fill = yellow_fill
        led.cell(r, 7, round(e.amount, 2)).number_format = inr
        led.cell(r, 7).fill = yellow_fill
        led.cell(r, 8, e.from_acct).fill = yellow_fill
        led.cell(r, 9, e.to_acct).fill = yellow_fill
        led.cell(r, 10, e.category).fill = yellow_fill
        led.cell(r, 11, e.budget).fill = yellow_fill
        led.cell(r, 12, e.recurring).fill = yellow_fill
        led.cell(r, 13, e.notes).fill = yellow_fill
        for c in range(1, 14):
            led.cell(r, c).border = thin

    # Pre-fill formulas for extra empty rows
    for r in range(2 + n, 2 + capacity):
        led.cell(r, 3, f'=IF(A{r}="","",TEXT(A{r},"dddd"))')
        led.cell(r, 4, f'=IF(A{r}="","",TEXT(A{r},"MMMM"))')
        led.cell(r, 5, f'=IF(A{r}="","",YEAR(A{r}))')
        for c in [1, 2, 6, 7, 8, 9, 10, 11, 12, 13]:
            led.cell(r, c).fill = yellow_fill
            led.cell(r, c).border = thin
        led.cell(r, 7).number_format = inr
        led.cell(r, 1).number_format = "dd/mm/yyyy"
        led.cell(r, 2).number_format = "HH:mm"

    led.auto_filter.ref = f"A1:M{1 + capacity}"

    # Reconciliation actuals (closing as of statements)
    rec = wb["Reconciliation"]
    # rows 5–15 = accounts in config order
    actuals = {
        5: 47649.45,   # HDFC Savings
        6: 142.11,     # ICICI Savings
        9: None,       # HDFC CC due — leave blank (partial cycle)
        10: 3795.04,   # ICICI CC due estimate
    }
    for row, val in actuals.items():
        if val is not None:
            rec.cell(row, 4, val).number_format = inr
            rec.cell(row, 4).fill = yellow_fill
            rec.cell(row, 6, date(2026, 8, 6))
            rec.cell(row, 6).number_format = "dd/mm/yyyy"

    out = WORKBOOK
    wb.save(out)

    # verification
    report = {
        "n_entries": n,
        "by_source": {},
        "by_type": {},
        "balances": {},
        "targets": TARGETS,
        "rent_rows": [],
        "icici_cc_pay_rows": [],
    }
    for e in entries:
        report["by_source"][e.source] = report["by_source"].get(e.source, 0) + 1
        report["by_type"][e.typ] = report["by_type"].get(e.typ, 0) + 1
        if e.category == "Rent":
            report["rent_rows"].append((str(e.date), e.amount, e.notes[:60]))
        if e.typ == "Credit Card Payment" and "ICICI" in e.to_acct:
            report["icici_cc_pay_rows"].append((str(e.date), e.amount, e.from_acct, e.notes[:60]))

    for name, liab in [
        ("HDFC Savings", False),
        ("ICICI Savings", False),
        ("HDFC Credit Card", True),
        ("ICICI Credit Card", True),
    ]:
        report["balances"][name] = round(calc_balance(name, entries, liability=liab), 2)

    return report


def main():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        entries = build_entries(tmp)
        report = write_workbook(entries)

    print("=== IMPORT REPORT ===")
    print(f"Entries written: {report['n_entries']}")
    print("By source:", report["by_source"])
    print("By type:", report["by_type"])
    print("\nRent rows:")
    for r in report["rent_rows"]:
        print(" ", r)
    print("\nICICI CC payment rows:")
    for r in report["icici_cc_pay_rows"]:
        print(" ", r)
    print("\nCalculated balances vs targets:")
    for name, bal in report["balances"].items():
        tgt = report["targets"].get(name)
        if tgt is not None:
            diff = round(bal - tgt, 2)
            ok = "OK" if abs(diff) < 0.05 else f"DIFF {diff}"
            print(f"  {name}: calc={bal} target={tgt} {ok}")
        else:
            print(f"  {name}: calc={bal}")
    print(f"\nWorkbook: {WORKBOOK}")
    print(f"Backup:   {BACKUP}")


if __name__ == "__main__":
    main()
