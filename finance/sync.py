"""
Finance tracker — local entry log + LibreOffice workbook.

Phone form entries are validated, appended to data/finance/entries.jsonl,
and written into the Ledger tab of the live finance workbook.
"""

from __future__ import annotations

import json
import os
import threading
import uuid
from calendar import monthrange
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
FINANCE_DIR = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data" / "finance"
ENTRIES_PATH = DATA_DIR / "entries.jsonl"
# Live workbook lives under Documents (not the repo copy).
DEFAULT_WORKBOOK = Path("/home/himanshu/Documents/Finance/Finance-Mng-V2.xlsx")

DEFAULT_TZ = "Asia/Kolkata"

_yellow = PatternFill("solid", fgColor="FFF2CC")
_thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_inr = '₹#,##0.00'
_workbook_lock = threading.Lock()

TYPES = [
    "Expense",
    "Income",
    "Transfer",
    "Credit Card Payment",
    "Refund",
    "Investment",
    "Adjustment",
]

# Accept old type name from earlier form versions
TYPE_ALIASES = {
    "Credit Card Bill Payment": "Credit Card Payment",
}

# Fallback if workbook is missing; live options prefer Configuration sheet.
CATEGORIES = [
    # Food
    "Groceries - Online",
    "Groceries - Physical",
    "Eating outside",
    "Food delivery",
    "Cafe / Snacks",
    # Transport
    "Petrol",
    "Cab / Auto",
    "Metro / Bus",
    "Parking",
    "Vehicle service",
    # Home
    "House cleaning (UC)",
    "Electricity bill",
    "Internet / WiFi",
    "Water / Gas",
    "Home maintenance",
    "Household / Kitchen",
    "Rent",
    # Personal
    "Salon",
    "Pharmacy / Medicine",
    "Medical",
    "Fitness / Gym",
    "Clothes / Fashion",
    "Personal care",
    # Lifestyle
    "Subscription",
    "Subscription - OTT",
    "Subscription - Software",
    "Entertainment",
    "Shopping - Online",
    "Shopping - Offline",
    "Gifts",
    "Education / Courses",
    "Mobile recharge",
    "Laptop / Electronics",
    # Travel
    "Travel / Flight",
    "Hotels / Stay",
    "Travel - Local",
    # Fixed
    "EMIs",
    "Insurance",
    "Bank Charges",
    "Credit Card Bill",
    # Income
    "Salary",
    "Cashback",
    "Refund",
    "Reimbursement",
    "Interest / Dividends",
    "Bonus",
    # Finance
    "FD Deposit",
    "FD Maturity",
    "Investment",
    "Transfer",
    "Reconciliation",
    # Catch-all
    "Other",
]

ACCOUNTS = [
    "HDFC Savings",
    "ICICI Savings",
    "Cash",
    "Wallet",
    "HDFC Credit Card",
    "ICICI Credit Card",
    "FD",
    "Mutual Fund",
    "Employer",
    "Expense",
    "External",
]

# Types that usually should not count toward monthly budget
NON_BUDGET_TYPES = {
    "Income",
    "Transfer",
    "Credit Card Payment",
    "Investment",
    "Adjustment",
}

DEFAULTS = {
    "type": "Expense",
    "category": "",
    "from_account": "HDFC Savings",
    "to_account": "Expense",
    "include_in_budget": True,
    "notes": "",
    "source": "manual",
}

# Who wrote the Ledger row (column M). Form = manual; voice LLM = ai.
SOURCES = ("manual", "ai")

# Sensible From/To when the client only sends type (or type changes)
TYPE_ACCOUNT_DEFAULTS: dict[str, tuple[str, str, bool]] = {
    # type → (from, to, include_in_budget)
    "Expense": ("HDFC Savings", "Expense", True),
    "Income": ("Employer", "HDFC Savings", False),
    "Transfer": ("HDFC Savings", "ICICI Savings", False),
    "Credit Card Payment": ("HDFC Savings", "HDFC Credit Card", False),
    "Refund": ("Expense", "HDFC Savings", True),
    "Investment": ("HDFC Savings", "Mutual Fund", False),
    "Adjustment": ("HDFC Savings", "Expense", False),
}


def _tz() -> ZoneInfo:
    name = os.environ.get("FINANCE_TZ", DEFAULT_TZ).strip() or DEFAULT_TZ
    try:
        return ZoneInfo(name)
    except Exception:
        return ZoneInfo(DEFAULT_TZ)


def today_iso() -> str:
    return datetime.now(_tz()).date().isoformat()


def now_time_str() -> str:
    """Local wall-clock time (24h HH:mm)."""
    return datetime.now(_tz()).strftime("%H:%M")


def workbook_path() -> Path:
    override = os.environ.get("FINANCE_WORKBOOK", "").strip()
    return Path(override) if override else DEFAULT_WORKBOOK


def _read_config_lists(path: Path) -> tuple[list[str], list[str]]:
    """Read categories + accounts from Configuration sheet when present."""
    cats: list[str] = []
    accs: list[str] = []
    if not path.is_file():
        return cats, accs
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return cats, accs
    try:
        if "Configuration" not in wb.sheetnames:
            return cats, accs
        cfg = wb["Configuration"]
        # Accounts: column A starting row 11 until blank (Type in B)
        for r in range(11, 80):
            name = cfg.cell(r, 1).value
            typ = cfg.cell(r, 2).value
            if not name or not typ:
                break
            accs.append(str(name).strip())
        # Categories: find header "Category" then names in col A (skip Group header row)
        cat_start = None
        for r in range(1, 120):
            v = cfg.cell(r, 1).value
            if v == "Category" and cfg.cell(r, 2).value in ("Group", "Typical Budget?"):
                cat_start = r + 1
                break
            if v == "Category" and str(cfg.cell(r, 2).value or "").startswith("Typical"):
                cat_start = r + 1
                break
        if cat_start is None:
            # Fallback: old layout Category at row 25
            for r in range(24, 40):
                if cfg.cell(r, 1).value == "Category":
                    cat_start = r + 1
                    break
        if cat_start:
            for r in range(cat_start, cat_start + 120):
                name = cfg.cell(r, 1).value
                if not name:
                    # allow blank gaps only if further filled? stop on first blank
                    break
                s = str(name).strip()
                if s in {"RULES", "Category"}:
                    break
                cats.append(s)
    finally:
        wb.close()
    return cats, accs


def options_payload() -> dict[str, Any]:
    wb = workbook_path()
    sheet_cats, sheet_accs = _read_config_lists(wb)
    return {
        "types": TYPES,
        "categories": sheet_cats or CATEGORIES,
        "accounts": sheet_accs or ACCOUNTS,
        "type_account_defaults": {
            t: {"from_account": f, "to_account": to, "include_in_budget": bud}
            for t, (f, to, bud) in TYPE_ACCOUNT_DEFAULTS.items()
        },
        "defaults": DEFAULTS,
        "today": today_iso(),
        "timezone": str(_tz()),
        "workbook": str(wb),
        "workbook_exists": wb.is_file(),
    }


def _allowed_accounts() -> set[str]:
    """Accounts allowed for validation: Configuration sheet ∪ defaults."""
    try:
        _, sheet_accs = _read_config_lists(workbook_path())
    except Exception:
        sheet_accs = []
    return set(sheet_accs or []) | set(ACCOUNTS)


def _coerce_bool(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return bool(value)


def _map_legacy_accounts(raw: dict[str, Any], t: str) -> tuple[str, str]:
    """payment_method + credit_card → from/to (old clients)."""
    pay = str(raw.get("payment_method") or raw.get("paymentMethod") or "").strip()
    card = str(raw.get("credit_card") or raw.get("creditCard") or "-").strip()
    card_account = (
        "ICICI Credit Card"
        if card == "ICICI"
        else "HDFC Credit Card"
    )
    savings = "Cash" if pay == "Cash" else "HDFC Savings"

    if t == "Income":
        return "Employer", savings
    if t == "Refund":
        if pay == "Credit Card":
            return "Expense", card_account
        return "Expense", savings
    if t == "Investment":
        return savings, "Mutual Fund"
    if t == "Transfer":
        return savings, "ICICI Savings"
    if t == "Credit Card Payment":
        return savings, card_account
    if t == "Adjustment":
        return savings, "Expense"
    # Expense
    if pay == "Credit Card":
        return card_account, "Expense"
    if pay == "Cash":
        return "Cash", "Expense"
    return savings, "Expense"


def _normalize_entry(raw: dict[str, Any]) -> dict[str, Any]:
    amount_raw = raw.get("amount")
    try:
        amount = float(amount_raw)
    except (TypeError, ValueError) as e:
        raise ValueError("Amount must be a number") from e
    if amount <= 0:
        raise ValueError("Amount must be greater than 0")
    if amount > 1e9:
        raise ValueError("Amount too large")

    t = str(raw.get("type") or DEFAULTS["type"]).strip()
    t = TYPE_ALIASES.get(t, t)
    if t not in TYPES:
        raise ValueError(f"Invalid type: {t}")

    category = str(raw.get("category") or "").strip()
    if not category:
        raise ValueError("Category is required")
    if len(category) > 80:
        raise ValueError("Category too long")

    from_account = str(
        raw.get("from_account") or raw.get("fromAccount") or ""
    ).strip()
    to_account = str(raw.get("to_account") or raw.get("toAccount") or "").strip()

    if not from_account or not to_account:
        # Prefer type defaults, then legacy payment fields
        if raw.get("payment_method") or raw.get("credit_card"):
            lf, lt = _map_legacy_accounts(raw, t)
            from_account = from_account or lf
            to_account = to_account or lt
        else:
            df, dt, _ = TYPE_ACCOUNT_DEFAULTS.get(
                t, (DEFAULTS["from_account"], DEFAULTS["to_account"], True)
            )
            from_account = from_account or df
            to_account = to_account or dt

    allowed_accs = _allowed_accounts()
    if from_account not in allowed_accs:
        raise ValueError(f"Invalid from_account: {from_account}")
    if to_account not in allowed_accs:
        raise ValueError(f"Invalid to_account: {to_account}")
    if from_account == to_account:
        raise ValueError("from_account and to_account must differ")

    default_budget = TYPE_ACCOUNT_DEFAULTS.get(t, (None, None, True))[2]
    if "include_in_budget" in raw:
        include = _coerce_bool(raw.get("include_in_budget"), default_budget)
    else:
        include = False if t in NON_BUDGET_TYPES else default_budget

    notes = str(raw.get("notes") or "").strip()
    if len(notes) > 500:
        raise ValueError("Notes too long (max 500)")

    source = str(raw.get("source") or DEFAULTS["source"]).strip().lower()
    if source not in SOURCES:
        raise ValueError(f"Invalid source: {source} (use manual or ai)")

    note_id = str(raw.get("note_id") or raw.get("noteId") or "").strip() or None
    if note_id and len(note_id) > 40:
        raise ValueError("note_id too long")

    now = datetime.now(_tz())

    # Optional override date (YYYY-MM-DD); default today in FINANCE_TZ
    date_str = str(raw.get("date") or "").strip() or now.date().isoformat()
    try:
        date.fromisoformat(date_str)
    except ValueError as e:
        raise ValueError("Date must be YYYY-MM-DD") from e

    # Optional override time (HH:mm or HH:mm:ss); default now in FINANCE_TZ
    time_str = str(raw.get("time") or "").strip()
    if time_str:
        parts = time_str.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            s = int(parts[2]) if len(parts) > 2 else 0
            if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= s <= 59):
                raise ValueError
            time_str = f"{h:02d}:{m:02d}" if s == 0 else f"{h:02d}:{m:02d}:{s:02d}"
        except (ValueError, IndexError) as e:
            raise ValueError("Time must be HH:mm") from e
    else:
        time_str = now.strftime("%H:%M")

    entry = {
        "id": uuid.uuid4().hex[:12],
        "date": date_str,
        "time": time_str,
        "type": t,
        "amount": amount,
        "category": category,
        "from_account": from_account,
        "to_account": to_account,
        "include_in_budget": include,
        "notes": notes,
        "source": source,
        "created_at": now.replace(microsecond=0).isoformat(),
    }
    if note_id:
        entry["note_id"] = note_id
    return entry


def _append_local(entry: dict[str, Any]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    line = json.dumps(entry, ensure_ascii=False) + "\n"
    with open(ENTRIES_PATH, "a", encoding="utf-8") as f:
        f.write(line)
        f.flush()
        os.fsync(f.fileno())


def _parse_time(time_str: str) -> time | None:
    s = (time_str or "").strip()
    if not s:
        return None
    parts = s.split(":")
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        sec = int(parts[2]) if len(parts) > 2 else 0
        return time(h, m, sec)
    except (ValueError, IndexError):
        return None


def _last_ledger_data_row(led) -> int:
    """Highest row that already has a Date or Type value."""
    last = 1
    for r in range(2, (led.max_row or 1) + 1):
        if led.cell(r, 1).value is not None or led.cell(r, 6).value is not None:
            last = r
    return last


def _entry_accounts(entry: dict[str, Any]) -> tuple[str, str]:
    """Resolve From/To, including legacy payment_method rows in the JSONL log."""
    fa = str(entry.get("from_account") or entry.get("fromAccount") or "").strip()
    ta = str(entry.get("to_account") or entry.get("toAccount") or "").strip()
    if fa and ta:
        return fa, ta
    t = str(entry.get("type") or "Expense").strip()
    t = TYPE_ALIASES.get(t, t)
    if entry.get("payment_method") or entry.get("credit_card"):
        return _map_legacy_accounts(entry, t)
    df, dt, _ = TYPE_ACCOUNT_DEFAULTS.get(
        t, (DEFAULTS["from_account"], DEFAULTS["to_account"], True)
    )
    return fa or df, ta or dt


def _ensure_source_header(led) -> None:
    """Ensure column M header is 'Source' (surgical — header only)."""
    hdr = led.cell(1, 13).value
    if hdr is None or str(hdr).strip() == "":
        led.cell(1, 13, "Source")
        # Match other header cells lightly if they have fill
        try:
            led.cell(1, 13).fill = led.cell(1, 12).fill
            led.cell(1, 13).font = led.cell(1, 12).font
            led.cell(1, 13).border = led.cell(1, 12).border
        except Exception:
            pass
    elif str(hdr).strip().lower() not in {"source", "added by", "added_by", "who"}:
        # Do not clobber an unexpected column M — still write data as Source label
        # only when empty; if user repurposed M, keep writing source in M for our rows
        # but leave the existing header text alone.
        pass


def _fill_ledger_row(
    led,
    r: int,
    entry: dict[str, Any],
    *,
    preserve_source: str | None = None,
) -> None:
    """Write one Ledger row (cols A–M). Day/Month/Year get formulas."""
    d = date.fromisoformat(str(entry["date"])[:10])
    tval = _parse_time(str(entry.get("time") or ""))
    from_account, to_account = _entry_accounts(entry)
    if preserve_source is not None:
        source = str(preserve_source).strip()
    else:
        source = str(entry.get("source") or "manual").strip().lower()
        if source not in SOURCES:
            source = "manual"

    led.cell(r, 1, d).number_format = "dd/mm/yyyy"
    led.cell(r, 1).fill = _yellow

    if tval is not None:
        led.cell(r, 2, tval).number_format = "HH:mm"
    else:
        led.cell(r, 2, None).number_format = "HH:mm"
    led.cell(r, 2).fill = _yellow

    # Day / Month / Year formulas (match existing workbook template)
    led.cell(r, 3, f'=IF(A{r}="","",TEXT(A{r},"dddd"))')
    led.cell(r, 4, f'=IF(A{r}="","",TEXT(A{r},"MMMM"))')
    led.cell(r, 5, f'=IF(A{r}="","",YEAR(A{r}))')

    led.cell(r, 6, entry["type"]).fill = _yellow
    led.cell(r, 7, round(float(entry["amount"]), 2)).number_format = _inr
    led.cell(r, 7).fill = _yellow
    led.cell(r, 8, from_account).fill = _yellow
    led.cell(r, 9, to_account).fill = _yellow
    led.cell(r, 10, entry["category"]).fill = _yellow
    # Match existing rows: =TRUE() / =FALSE() so SUMIFS(...TRUE()) works
    led.cell(r, 11, "=TRUE()" if entry.get("include_in_budget") else "=FALSE()")
    led.cell(r, 11).fill = _yellow
    led.cell(r, 12, entry.get("notes") or "").fill = _yellow
    led.cell(r, 13, source).fill = _yellow

    for c in range(1, 14):
        led.cell(r, c).border = _thin


def _extend_autofilter(led, last_row: int) -> None:
    end = max(last_row, 50)
    # Keep room for future empty formula rows
    end = max(end, led.max_row or end)
    led.auto_filter.ref = f"A1:M{end}"


def append_to_workbook(entry: dict[str, Any]) -> dict[str, Any]:
    """
    Append one entry to the Ledger sheet of the live finance workbook.

    Safe to call while the phone form is in use. If LibreOffice has the
    file open for exclusive write, save may fail — close LO and retry, or
    re-open the file after the save (LO needs Reload to see external writes).
    """
    path = workbook_path()
    if not path.is_file():
        return {
            "ok": False,
            "error": f"Workbook not found: {path}",
        }

    with _workbook_lock:
        try:
            wb = load_workbook(path)
        except Exception as e:
            return {"ok": False, "error": f"Could not open workbook: {e}"}

        if "Ledger" not in wb.sheetnames:
            return {"ok": False, "error": "Workbook has no Ledger sheet"}

        led = wb["Ledger"]
        _ensure_source_header(led)
        r = _last_ledger_data_row(led) + 1
        _fill_ledger_row(led, r, entry)
        _extend_autofilter(led, r)

        tmp = path.with_suffix(".xlsx.writing")
        try:
            wb.save(tmp)
            os.replace(tmp, path)
        except PermissionError:
            try:
                if tmp.is_file():
                    tmp.unlink()
            except OSError:
                pass
            return {
                "ok": False,
                "error": (
                    "Cannot write workbook (permission denied). "
                    "Close LibreOffice if Finance-Mng-V2.xlsx is open, then retry."
                ),
            }
        except Exception as e:
            try:
                if tmp.is_file():
                    tmp.unlink()
            except OSError:
                pass
            return {"ok": False, "error": f"Workbook save failed: {e}"}

        return {
            "ok": True,
            "path": str(path),
            "sheet": "Ledger",
            "row": r,
        }


def add_entry(raw: dict[str, Any]) -> dict[str, Any]:
    """Validate, log to JSONL, and append to the local xlsx Ledger."""
    entry = _normalize_entry(raw)
    _append_local(entry)
    sheet = append_to_workbook(entry)
    entry["sheet"] = sheet
    if sheet.get("ok"):
        entry["status"] = "saved"
    else:
        entry["status"] = "log_only"
    return entry


def recent_entries(limit: int = 20) -> list[dict[str, Any]]:
    if not ENTRIES_PATH.is_file():
        return []
    rows: list[dict[str, Any]] = []
    try:
        with open(ENTRIES_PATH, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    except OSError:
        return []
    return list(reversed(rows[-limit:]))


def _as_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _as_budget_flag(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    s = str(value).strip().lower()
    # Live sheet stores =TRUE() / =FALSE() formulas (not evaluated by openpyxl)
    if s in {"1", "true", "yes", "on", "=true()", "=true"}:
        return True
    if s in {"0", "false", "no", "off", "=false()", "=false"}:
        return False
    return False


def _month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def _add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    return date(y, m, 1)


def _account_balance(
    opening: float,
    acc_type: str,
    from_sum: float,
    to_sum: float,
) -> float:
    """Match Reconciliation calculated balance formulas."""
    if acc_type == "Liability":
        return opening + from_sum - to_sum
    return opening + to_sum - from_sum


def _budget_for_month(mb_grid: list[tuple[date, float]], target: date, default: float) -> float:
    for start, amount in mb_grid:
        if start.year == target.year and start.month == target.month:
            return amount
    return default


def _chart_category_bucket(typ: str, category: str) -> tuple[str, str]:
    """
    Map a ledger row to a mobile chart bucket.

    Income is excluded by the caller. Refund type and Refund/Reimbursement
    categories collapse into one credit bucket.
    """
    cat = (category or "").strip()
    t = (typ or "").strip()
    cat_l = cat.lower()
    if t == "Refund" or cat_l in ("refund", "reimbursement"):
        return "Refund / Reimbursement", "credit"
    if cat:
        return cat, "spend"
    if t:
        return t, "spend"
    return "Other", "spend"


def simple_dashboard_payload() -> dict[str, Any]:
    """
    Compute Simple Dashboard metrics from the live workbook.

    openpyxl cannot evaluate Excel formulas, so values are recomputed from
    Configuration + Ledger + Monthly Budget grid using the same rules as the sheet.
    """
    path = workbook_path()
    if not path.is_file():
        return {"ok": False, "error": f"Workbook not found: {path}"}

    today = datetime.now(_tz()).date()
    year = today.year
    # days left incl. today (matches EOMONTH(TODAY(),0)-TODAY()+1)
    days_in_month = monthrange(today.year, today.month)[1]
    days_left = days_in_month - today.day + 1
    month_elapsed_pct = today.day / days_in_month if days_in_month else 0.0

    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception as e:
        return {"ok": False, "error": f"Could not open workbook: {e}"}

    try:
        if "Ledger" not in wb.sheetnames:
            return {"ok": False, "error": "Workbook has no Ledger sheet"}
        if "Configuration" not in wb.sheetnames:
            return {"ok": False, "error": "Workbook has no Configuration sheet"}

        cfg = wb["Configuration"]
        led = wb["Ledger"]

        default_budget = _as_float(cfg.cell(6, 2).value, 31000.0)
        monthly_salary = _as_float(cfg.cell(7, 2).value, 0.0)

        # Accounts: rows 11+ until blank (name + type + opening)
        accounts: dict[str, dict[str, Any]] = {}
        liquid_names = ("HDFC Savings", "ICICI Savings", "Cash", "Wallet")
        for r in range(11, 80):
            name = cfg.cell(r, 1).value
            typ = cfg.cell(r, 2).value
            if not name or not typ:
                break
            name_s = str(name).strip()
            accounts[name_s] = {
                "type": str(typ).strip(),
                "opening": _as_float(cfg.cell(r, 3).value, 0.0),
            }

        # Monthly Budget grid (A20:B…) — only Budget column is manual
        mb_grid: list[tuple[date, float]] = []
        if "Monthly Budget" in wb.sheetnames:
            mb = wb["Monthly Budget"]
            for r in range(20, 80):
                raw = mb.cell(r, 1).value
                if raw is None:
                    break
                if isinstance(raw, datetime):
                    start = raw.date()
                elif isinstance(raw, date):
                    start = raw
                else:
                    continue
                mb_grid.append((start, _as_float(mb.cell(r, 2).value, default_budget)))

        this_month_start = _month_start(today)
        next_month_start = _add_months(this_month_start, 1)
        budget = _budget_for_month(mb_grid, this_month_start, default_budget)
        next_month_budget = _budget_for_month(mb_grid, next_month_start, default_budget)

        from_sums: dict[str, float] = {n: 0.0 for n in accounts}
        to_sums: dict[str, float] = {n: 0.0 for n in accounts}
        budget_expenses = 0.0
        budget_refunds = 0.0
        # Category charts: full month (all non-Income) vs budget-flagged only
        cat_totals_full: dict[str, float] = {}
        cat_kinds_full: dict[str, str] = {}  # spend | credit
        cat_totals_budget: dict[str, float] = {}
        cat_kinds_budget: dict[str, str] = {}

        for row in led.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            # A Date, B Time, C Day, D Month, E Year, F Type, G Amount,
            # H From, I To, J Category, K Budget, L Notes, M Source
            # Month/Year/Budget are often formulas — use Date for month filter.
            typ = str(row[5] or "").strip() if len(row) > 5 else ""
            amount = _as_float(row[6] if len(row) > 6 else None, 0.0)
            from_acc = str(row[7] or "").strip() if len(row) > 7 else ""
            to_acc = str(row[8] or "").strip() if len(row) > 8 else ""
            category = str(row[9] or "").strip() if len(row) > 9 else ""
            in_budget = _as_budget_flag(row[10] if len(row) > 10 else None)

            if from_acc in from_sums:
                from_sums[from_acc] += amount
            if to_acc in to_sums:
                to_sums[to_acc] += amount

            dval = row[0]
            if isinstance(dval, datetime):
                row_date = dval.date()
            elif isinstance(dval, date):
                row_date = dval
            else:
                continue
            if row_date.month != today.month or row_date.year != year:
                continue

            if typ == "Expense" and in_budget:
                budget_expenses += amount
            elif typ == "Refund" and in_budget:
                budget_refunds += amount

            # Chart buckets (exclude Income only)
            if typ != "Income" and amount != 0:
                cat_key, kind = _chart_category_bucket(typ, category)
                cat_totals_full[cat_key] = cat_totals_full.get(cat_key, 0.0) + amount
                if kind == "credit" or cat_key not in cat_kinds_full:
                    cat_kinds_full[cat_key] = kind
                if in_budget:
                    cat_totals_budget[cat_key] = (
                        cat_totals_budget.get(cat_key, 0.0) + amount
                    )
                    if kind == "credit" or cat_key not in cat_kinds_budget:
                        cat_kinds_budget[cat_key] = kind

        net_budget_spend = budget_expenses - budget_refunds
        budget_remaining = budget - net_budget_spend
        budget_used_pct = (net_budget_spend / budget) if budget else 0.0
        safe_per_day = (
            max(0.0, budget_remaining) / days_left if days_left > 0 else 0.0
        )

        if budget_remaining < 0:
            pace = "OVER BUDGET"
            pace_level = "danger"
        elif budget_used_pct > month_elapsed_pct + 0.05:
            pace = "Spending too fast"
            pace_level = "warn"
        elif budget_used_pct > month_elapsed_pct:
            pace = "Slightly ahead"
            pace_level = "warn"
        else:
            pace = "On track"
            pace_level = "ok"

        balances: dict[str, float] = {}
        for name, meta in accounts.items():
            balances[name] = _account_balance(
                meta["opening"],
                meta["type"],
                from_sums.get(name, 0.0),
                to_sums.get(name, 0.0),
            )

        liquid = sum(balances.get(n, 0.0) for n in liquid_names)
        hdfc_cc = balances.get("HDFC Credit Card", 0.0)
        icici_cc = balances.get("ICICI Credit Card", 0.0)
        cc_total = hdfc_cc + icici_cc
        budget_reserved = max(0.0, budget_remaining)
        free_to_allocate = liquid - budget_reserved
        est_free_next = free_to_allocate - cc_total + monthly_salary - next_month_budget

        def _cats_payload(
            totals: dict[str, float], kinds: dict[str, str]
        ) -> list[dict[str, Any]]:
            return [
                {
                    "name": name,
                    "amount": round(total, 2),
                    "kind": kinds.get(name, "spend"),
                }
                for name, total in sorted(
                    totals.items(),
                    key=lambda kv: (-abs(kv[1]), kv[0].lower()),
                )
            ]

        categories_full = _cats_payload(cat_totals_full, cat_kinds_full)
        categories_budget = _cats_payload(cat_totals_budget, cat_kinds_budget)
        cat_full_max = max(
            (abs(c["amount"]) for c in categories_full), default=0.0
        )
        cat_budget_max = max(
            (abs(c["amount"]) for c in categories_budget), default=0.0
        )
        # Back-compat: categories = full month (previous default)
        categories = categories_full
        cat_max = cat_full_max

        planned = _planned_expenses_summary(wb, today)

        return {
            "ok": True,
            "month": today.strftime("%B %Y"),
            "days_left": days_left,
            "pace": pace,
            "pace_level": pace_level,
            "budget": round(budget, 2),
            "budget_remaining": round(budget_remaining, 2),
            "budget_spent": round(net_budget_spend, 2),
            "safe_per_day": round(safe_per_day, 2),
            "budget_used_pct": round(budget_used_pct, 4),
            "month_elapsed_pct": round(month_elapsed_pct, 4),
            "hdfc_cc": round(hdfc_cc, 2),
            "icici_cc": round(icici_cc, 2),
            "cc_total": round(cc_total, 2),
            "liquid": round(liquid, 2),
            "budget_reserved": round(budget_reserved, 2),
            "free_to_allocate": round(free_to_allocate, 2),
            "monthly_salary": round(monthly_salary, 2),
            "next_month_budget": round(next_month_budget, 2),
            "est_free_next_month": round(est_free_next, 2),
            "categories": categories,
            "categories_max": round(cat_max, 2),
            "categories_full": categories_full,
            "categories_full_max": round(cat_full_max, 2),
            "categories_budget": categories_budget,
            "categories_budget_max": round(cat_budget_max, 2),
            "planned": planned,
        }
    finally:
        wb.close()


def _as_bool_flag(value: Any) -> bool:
    """TRUE / True / 1 / 'true' → True; blank/falsey → False."""
    if value is True or value == 1:
        return True
    if value is False or value is None or value == 0 or value == "":
        return False
    s = str(value).strip().lower()
    return s in ("true", "yes", "1", "y")


def _planned_cell_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel serial date (openpyxl usually already converts; keep safe fallback)
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s[:10])
        except ValueError:
            return None
    return None


def _planned_kind_column(pe) -> int | None:
    for c in range(1, 16):
        v = pe.cell(14, c).value
        if v and str(v).strip().lower() == "kind":
            return c
    return None


def _infer_recurring_kind(kind: Any, category: Any) -> str:
    k = str(kind or "").strip()
    if k:
        return k
    cat = str(category or "").strip()
    if cat == "EMIs":
        return "Loan / EMI"
    if cat == "Investment":
        return "Investment"
    return "Lifestyle"


def _month_first(d: date) -> date:
    return date(d.year, d.month, 1)


def _add_calendar_months(d: date, n: int) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)


def _recurring_due_in_month(
    *,
    freq: str,
    amount: float,
    start: date | None,
    end: date | None,
    month: date,
) -> float:
    """Cash due in `month` (any day; compared as calendar month)."""
    ms = _month_first(month)
    me = date(ms.year, ms.month, monthrange(ms.year, ms.month)[1])
    if start is not None and start > me:
        return 0.0
    if end is not None and end < ms:
        return 0.0
    if freq == "Monthly":
        return amount
    if freq == "Yearly":
        if start is None or start.month != ms.month:
            return 0.0
        if date(start.year, start.month, 1) > ms:
            return 0.0
        return amount
    return 0.0


def _planned_expenses_summary(wb, today: date) -> dict[str, Any]:
    """
    Read Planned Expenses sheet (planning only) and recompute summary metrics.
    Never writes Ledger / balances. Missing sheet → zeros + available=False.
    """
    empty_forecast = [
        {
            "month": _add_calendar_months(_month_first(today), i).isoformat(),
            "label": _add_calendar_months(_month_first(today), i).strftime("%b %Y"),
            "loan_emi": 0.0,
            "lifestyle": 0.0,
            "investment": 0.0,
            "total": 0.0,
        }
        for i in range(6)
    ]
    empty = {
        "available": False,
        "monthly_fixed_cost": 0.0,
        "monthly_recurring_count": 0,
        "yearly_commitments": 0.0,
        "upcoming_30_days": 0.0,
        "upcoming_90_days": 0.0,
        "total_planned_one_time": 0.0,
        "this_month_loan_emi": 0.0,
        "this_month_lifestyle": 0.0,
        "this_month_investment": 0.0,
        "this_month_recurring_total": 0.0,
        "forecast_6m": empty_forecast,
    }
    if "Planned Expenses" not in wb.sheetnames:
        return empty

    pe = wb["Planned Expenses"]
    kind_col = _planned_kind_column(pe)
    recurring_count = 0
    yearly_commit = 0.0
    months = [_add_calendar_months(_month_first(today), i) for i in range(6)]
    buckets = [
        {"loan_emi": 0.0, "lifestyle": 0.0, "investment": 0.0}
        for _ in months
    ]

    # Recurring block: rows 15–64
    # A Expense, B Category, C Frequency, D Amount, E Start, F End,
    # G Active, H Monthly Equivalent, … Kind (header row 14)
    for r in range(15, 65):
        expense = pe.cell(r, 1).value
        if expense is None or str(expense).strip() == "":
            continue
        freq = str(pe.cell(r, 3).value or "").strip()
        amount = _as_float(pe.cell(r, 4).value, 0.0)
        start = _planned_cell_date(pe.cell(r, 5).value)
        end = _planned_cell_date(pe.cell(r, 6).value)
        active = _as_bool_flag(pe.cell(r, 7).value)
        category = pe.cell(r, 2).value
        kind_val = pe.cell(r, kind_col).value if kind_col else None
        kind = _infer_recurring_kind(kind_val, category)
        if freq == "Yearly" and active:
            yearly_commit += amount
        if active:
            recurring_count += 1
            for i, month in enumerate(months):
                due = _recurring_due_in_month(
                    freq=freq,
                    amount=amount,
                    start=start,
                    end=end,
                    month=month,
                )
                if due <= 0:
                    continue
                if kind == "Loan / EMI":
                    buckets[i]["loan_emi"] += due
                elif kind == "Investment":
                    buckets[i]["investment"] += due
                else:
                    buckets[i]["lifestyle"] += due

    forecast_6m = []
    for month, b in zip(months, buckets):
        total = b["loan_emi"] + b["lifestyle"] + b["investment"]
        forecast_6m.append(
            {
                "month": month.isoformat(),
                "label": month.strftime("%b %Y"),
                "loan_emi": round(b["loan_emi"], 2),
                "lifestyle": round(b["lifestyle"], 2),
                "investment": round(b["investment"], 2),
                "total": round(total, 2),
            }
        )

    this = forecast_6m[0]
    monthly_fixed = this["total"]

    upcoming_30 = 0.0
    upcoming_90 = 0.0
    total_planned = 0.0
    # One-time: rows 68–97
    # A Expense, B Cat, C Expected Month, D Expected Date, E Amount,
    # F Priority, G Status, H Notes, I Effective Date (formula — recompute)
    for r in range(68, 98):
        expense = pe.cell(r, 1).value
        if expense is None or str(expense).strip() == "":
            continue
        amount = _as_float(pe.cell(r, 5).value, 0.0)
        status = str(pe.cell(r, 7).value or "").strip()
        if status != "Planned":
            continue
        total_planned += amount
        edate = _planned_cell_date(pe.cell(r, 4).value)
        emonth = _planned_cell_date(pe.cell(r, 3).value)
        effective = edate or emonth
        if effective is None:
            continue
        delta = (effective - today).days
        if 0 <= delta <= 30:
            upcoming_30 += amount
        if 0 <= delta <= 90:
            upcoming_90 += amount

    return {
        "available": True,
        "monthly_fixed_cost": round(monthly_fixed, 2),
        "monthly_recurring_count": recurring_count,
        "yearly_commitments": round(yearly_commit, 2),
        "upcoming_30_days": round(upcoming_30, 2),
        "upcoming_90_days": round(upcoming_90, 2),
        "total_planned_one_time": round(total_planned, 2),
        "this_month_loan_emi": this["loan_emi"],
        "this_month_lifestyle": this["lifestyle"],
        "this_month_investment": this["investment"],
        "this_month_recurring_total": this["total"],
        "forecast_6m": forecast_6m,
    }


def _ledger_cell_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s[:10])
        except ValueError:
            return None
    return None


def _ledger_cell_time(value: Any) -> str:
    """Return HH:mm or empty string."""
    if value is None or value == "":
        return ""
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        # Excel serial fraction of a day
        try:
            total = int(round(float(value) * 24 * 60)) % (24 * 60)
            return f"{total // 60:02d}:{total % 60:02d}"
        except (TypeError, ValueError, OverflowError):
            return ""
    s = str(value).strip()
    if not s:
        return ""
    # Accept HH:MM or HH:MM:SS
    parts = s.replace(".", ":").split(":")
    if len(parts) >= 2:
        try:
            h, m = int(parts[0]), int(parts[1])
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    return s[:5] if len(s) >= 5 else s


def month_transactions_payload(
    year: int | None = None,
    month: int | None = None,
) -> dict[str, Any]:
    """
    Read-only: this month's Ledger rows for the phone UI.

    Content fields only (no Source / app metadata). Day/Month/Year formula
    columns are derived from Date so mobile can show full row meaning.
    Newest first.
    """
    path = workbook_path()
    if not path.is_file():
        return {"ok": False, "error": f"Workbook not found: {path}"}

    today = datetime.now(_tz()).date()
    y = int(year) if year else today.year
    m = int(month) if month else today.month
    if m < 1 or m > 12:
        return {"ok": False, "error": "month must be 1–12"}
    if y < 2000 or y > 2100:
        return {"ok": False, "error": "year out of range"}

    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception as e:
        return {"ok": False, "error": f"Could not open workbook: {e}"}

    try:
        if "Ledger" not in wb.sheetnames:
            return {"ok": False, "error": "Workbook has no Ledger sheet"}
        led = wb["Ledger"]
        items: list[dict[str, Any]] = []

        for r_idx, row in enumerate(led.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            row_date = _ledger_cell_date(row[0])
            if row_date is None:
                continue
            if row_date.year != y or row_date.month != m:
                continue

            typ = str(row[5] or "").strip() if len(row) > 5 else ""
            amount = _as_float(row[6] if len(row) > 6 else None, 0.0)
            from_acc = str(row[7] or "").strip() if len(row) > 7 else ""
            to_acc = str(row[8] or "").strip() if len(row) > 8 else ""
            category = str(row[9] or "").strip() if len(row) > 9 else ""
            in_budget = _as_budget_flag(row[10] if len(row) > 10 else None)
            notes = str(row[11] or "").strip() if len(row) > 11 else ""
            time_str = _ledger_cell_time(row[1] if len(row) > 1 else None)

            items.append(
                {
                    "row": r_idx,
                    "date": row_date.isoformat(),
                    "time": time_str,
                    "day": row_date.strftime("%A"),
                    "month": row_date.strftime("%B"),
                    "year": row_date.year,
                    "type": typ,
                    "amount": round(amount, 2),
                    "from_account": from_acc,
                    "to_account": to_acc,
                    "category": category,
                    "include_in_budget": in_budget,
                    "notes": notes,
                }
            )

        # Newest first (date desc, time desc, row desc)
        def _sort_key(it: dict[str, Any]) -> tuple:
            t = it.get("time") or ""
            return (it.get("date") or "", t, it.get("row") or 0)

        items.sort(key=_sort_key, reverse=True)
        month_label = date(y, m, 1).strftime("%B %Y")
        prev_m = m - 1
        prev_y = y
        if prev_m < 1:
            prev_m = 12
            prev_y -= 1
        next_m = m + 1
        next_y = y
        if next_m > 12:
            next_m = 1
            next_y += 1
        return {
            "ok": True,
            "year": y,
            "month": m,
            "month_label": month_label,
            "count": len(items),
            "transactions": items,
            "prev": {"year": prev_y, "month": prev_m},
            "next": {"year": next_y, "month": next_m},
            "is_current": y == today.year and m == today.month,
            "today": today.isoformat(),
        }
    finally:
        wb.close()


def _row_to_transaction(r_idx: int, row: tuple[Any, ...]) -> dict[str, Any] | None:
    """Parse one Ledger values_only tuple into a transaction dict."""
    if not row or row[0] is None:
        return None
    row_date = _ledger_cell_date(row[0])
    if row_date is None:
        return None
    typ = str(row[5] or "").strip() if len(row) > 5 else ""
    amount = _as_float(row[6] if len(row) > 6 else None, 0.0)
    from_acc = str(row[7] or "").strip() if len(row) > 7 else ""
    to_acc = str(row[8] or "").strip() if len(row) > 8 else ""
    category = str(row[9] or "").strip() if len(row) > 9 else ""
    in_budget = _as_budget_flag(row[10] if len(row) > 10 else None)
    notes = str(row[11] or "").strip() if len(row) > 11 else ""
    time_str = _ledger_cell_time(row[1] if len(row) > 1 else None)
    source_raw = row[12] if len(row) > 12 else None
    source = str(source_raw or "").strip().lower() if source_raw not in (None, "") else ""
    return {
        "row": r_idx,
        "date": row_date.isoformat(),
        "time": time_str,
        "day": row_date.strftime("%A"),
        "month": row_date.strftime("%B"),
        "year": row_date.year,
        "type": typ,
        "amount": round(amount, 2),
        "from_account": from_acc,
        "to_account": to_acc,
        "category": category,
        "include_in_budget": in_budget,
        "notes": notes,
        # source kept for internal update preserve; UI may ignore
        "source": source or None,
    }


def get_ledger_transaction(row: int) -> dict[str, Any]:
    """Read one Ledger row by 1-based sheet row number."""
    path = workbook_path()
    if not path.is_file():
        return {"ok": False, "error": f"Workbook not found: {path}"}
    r = int(row)
    if r < 2:
        return {"ok": False, "error": "Invalid row"}

    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception as e:
        return {"ok": False, "error": f"Could not open workbook: {e}"}

    try:
        if "Ledger" not in wb.sheetnames:
            return {"ok": False, "error": "Workbook has no Ledger sheet"}
        led = wb["Ledger"]
        max_r = led.max_row or 1
        if r > max_r:
            return {"ok": False, "error": f"Row {r} not found"}
        values = tuple(led.cell(r, c).value for c in range(1, 14))
        tx = _row_to_transaction(r, values)
        if not tx:
            return {"ok": False, "error": f"Row {r} is empty"}
        return {"ok": True, "transaction": tx}
    finally:
        wb.close()


def _save_workbook_atomic(wb, path: Path) -> dict[str, Any]:
    """Save workbook via temp file; return ok/error dict."""
    tmp = path.with_suffix(".xlsx.writing")
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    except PermissionError:
        try:
            if tmp.is_file():
                tmp.unlink()
        except OSError:
            pass
        return {
            "ok": False,
            "error": (
                "Cannot write workbook (permission denied). "
                "Close LibreOffice if Finance-Mng-V2.xlsx is open, then retry."
            ),
        }
    except Exception as e:
        try:
            if tmp.is_file():
                tmp.unlink()
        except OSError:
            pass
        return {"ok": False, "error": f"Workbook save failed: {e}"}
    return {"ok": True, "path": str(path)}


def update_ledger_row(row: int, raw: dict[str, Any]) -> dict[str, Any]:
    """
    Surgically rewrite one Ledger data row (content fields only).
    Preserves existing Source unless raw explicitly sets source.
    """
    r = int(row)
    if r < 2:
        raise ValueError("Invalid row")

    # Normalize without inventing a new ledger append identity
    entry = _normalize_entry(raw)
    # Explicit empty time on edit should clear the cell (not default to now)
    if "time" in raw and not str(raw.get("time") or "").strip():
        entry["time"] = ""
    path = workbook_path()
    if not path.is_file():
        return {"ok": False, "error": f"Workbook not found: {path}", "entry": entry}

    with _workbook_lock:
        try:
            wb = load_workbook(path)
        except Exception as e:
            return {"ok": False, "error": f"Could not open workbook: {e}", "entry": entry}

        if "Ledger" not in wb.sheetnames:
            return {"ok": False, "error": "Workbook has no Ledger sheet", "entry": entry}

        led = wb["Ledger"]
        last = _last_ledger_data_row(led)
        if r > last:
            return {"ok": False, "error": f"Row {r} not found", "entry": entry}

        existing_date = led.cell(r, 1).value
        if existing_date is None and led.cell(r, 6).value is None:
            return {"ok": False, "error": f"Row {r} is empty", "entry": entry}

        existing_source = led.cell(r, 13).value
        preserve: str | None
        if "source" in raw and raw.get("source") is not None and str(raw.get("source")).strip():
            preserve = None  # use normalized entry source
        else:
            preserve = (
                str(existing_source).strip()
                if existing_source not in (None, "")
                else (entry.get("source") or "manual")
            )

        _ensure_source_header(led)
        _fill_ledger_row(led, r, entry, preserve_source=preserve)
        _extend_autofilter(led, last)

        saved = _save_workbook_atomic(wb, path)
        if not saved.get("ok"):
            return {
                "ok": False,
                "error": saved.get("error") or "save failed",
                "entry": entry,
            }

        log_entry = {
            **entry,
            "action": "update",
            "row": r,
            "source": preserve if preserve is not None else entry.get("source"),
        }
        try:
            _append_local(log_entry)
        except OSError:
            pass

        return {
            "ok": True,
            "row": r,
            "entry": entry,
            "path": str(path),
            "sheet": "Ledger",
            "status": "updated",
        }


def delete_ledger_row(row: int) -> dict[str, Any]:
    """Surgically delete one Ledger data row (shifts rows up)."""
    r = int(row)
    if r < 2:
        raise ValueError("Invalid row")

    path = workbook_path()
    if not path.is_file():
        return {"ok": False, "error": f"Workbook not found: {path}"}

    with _workbook_lock:
        try:
            wb = load_workbook(path)
        except Exception as e:
            return {"ok": False, "error": f"Could not open workbook: {e}"}

        if "Ledger" not in wb.sheetnames:
            return {"ok": False, "error": "Workbook has no Ledger sheet"}

        led = wb["Ledger"]
        last = _last_ledger_data_row(led)
        if r > last:
            return {"ok": False, "error": f"Row {r} not found"}

        # Snapshot content for the JSONL audit log
        values = tuple(led.cell(r, c).value for c in range(1, 14))
        snapshot = _row_to_transaction(r, values)

        if led.cell(r, 1).value is None and led.cell(r, 6).value is None:
            return {"ok": False, "error": f"Row {r} is empty"}

        led.delete_rows(r, 1)

        # Day/Month/Year formulas must reference their new row numbers after shift
        new_last = _last_ledger_data_row(led)
        for rr in range(r, new_last + 1):
            if led.cell(rr, 1).value is None and led.cell(rr, 6).value is None:
                continue
            led.cell(rr, 3, f'=IF(A{rr}="","",TEXT(A{rr},"dddd"))')
            led.cell(rr, 4, f'=IF(A{rr}="","",TEXT(A{rr},"MMMM"))')
            led.cell(rr, 5, f'=IF(A{rr}="","",YEAR(A{rr}))')

        _extend_autofilter(led, max(new_last, 50))
        saved = _save_workbook_atomic(wb, path)
        if not saved.get("ok"):
            return {"ok": False, "error": saved.get("error") or "save failed"}

        try:
            _append_local(
                {
                    "id": uuid.uuid4().hex[:12],
                    "action": "delete",
                    "row": r,
                    "deleted": snapshot,
                    "created_at": datetime.now(_tz()).replace(microsecond=0).isoformat(),
                }
            )
        except OSError:
            pass

        return {
            "ok": True,
            "row": r,
            "status": "deleted",
            "path": str(path),
            "sheet": "Ledger",
            "deleted": snapshot,
        }
