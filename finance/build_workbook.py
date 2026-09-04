#!/usr/bin/env python3
"""
Rebuild finance/Finance Mng.xlsx as a ledger-based single source of truth.

Requires openpyxl:
  ./whisper/.venv/bin/python finance/build_workbook.py

AI / automation policy (see ../AGENTS.md):
  - Default sheet work is SURGICAL on the live Finance-Mng-V2.xlsx — edit only what was asked.
  - Do NOT run --patch-live for small dashboard tweaks; it recreates whole
    Simple + Detailed sheets and wipes LibreOffice layout/copy/chart edits.
  - --patch-live only when the user explicitly requests a full dashboard rebuild.
"""
from __future__ import annotations

from datetime import date as ddate
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "Finance Mng.xlsx"
# Live workbook under Documents (not the repo copy).
LIVE = Path("/home/himanshu/Documents/Finance/Finance-Mng-V2.xlsx")

title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
muted_font = Font(name="Calibri", size=10, italic=True, color="666666")
money_font = Font(name="Calibri", size=11)
big_num_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
warn_font = Font(name="Calibri", size=12, bold=True, color="B71C1C")

header_fill = PatternFill("solid", fgColor="1F4E79")
section_fill = PatternFill("solid", fgColor="D6EAF8")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
calc_fill = PatternFill("solid", fgColor="E8F5E9")
good_fill = PatternFill("solid", fgColor="C8E6C9")
alert_fill = PatternFill("solid", fgColor="FFCDD2")
soft_fill = PatternFill("solid", fgColor="FFF8E1")

thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
inr = "₹#,##0.00"
pct = "0.0%"

ACCOUNTS = [
    ("HDFC Savings", "Asset", 3851.96, None, True, "Primary savings (Aug-only start)"),
    ("ICICI Savings", "Asset", 142.11, None, True, ""),
    ("Cash", "Asset", 0, None, True, ""),
    ("Wallet", "Asset", 0, None, True, "UPI wallets etc."),
    ("HDFC Credit Card", "Liability", 0, 396000, True, "Opening due; Aug-only = this cycle spends"),
    ("ICICI Credit Card", "Liability", 11723.51, 100000, True, "Opening = due at ledger start"),
    ("FD", "Asset", 0, None, True, "Fixed deposits"),
    ("Mutual Fund", "Asset", 0, None, True, "Investments"),
    ("Employer", "Virtual", 0, None, False, "Counterparty for salary / income"),
    ("Expense", "Virtual", 0, None, False, "Counterparty for spends"),
    ("External", "Virtual", 0, None, False, "Outside world / unknown source"),
]

TYPES = [
    "Income",
    "Expense",
    "Transfer",
    "Credit Card Payment",
    "Refund",
    "Investment",
    "Adjustment",
]

# (name, group, typical_include_in_budget)
# Groups make filtering/review easier; keep old names so existing Ledger rows still match.
CATEGORIES = [
    # Food & drink
    ("Groceries - Online", "Food", True),
    ("Groceries - Physical", "Food", True),
    ("Eating outside", "Food", True),
    ("Food delivery", "Food", True),
    ("Cafe / Snacks", "Food", True),
    # Transport
    ("Petrol", "Transport", True),
    ("Cab / Auto", "Transport", True),
    ("Metro / Bus", "Transport", True),
    ("Parking", "Transport", True),
    ("Vehicle service", "Transport", False),
    # Home & utilities
    ("House cleaning (UC)", "Home", True),
    ("Electricity bill", "Home", True),
    ("Internet / WiFi", "Home", True),
    ("Water / Gas", "Home", True),
    ("Home maintenance", "Home", False),
    ("Household / Kitchen", "Home", True),
    ("Rent", "Home", False),
    # Personal
    ("Salon", "Personal", True),
    ("Pharmacy / Medicine", "Personal", True),
    ("Medical", "Personal", False),
    ("Fitness / Gym", "Personal", True),
    ("Clothes / Fashion", "Personal", True),
    ("Personal care", "Personal", True),
    # Lifestyle & shopping
    ("Subscription", "Lifestyle", True),
    ("Subscription - OTT", "Lifestyle", True),
    ("Subscription - Software", "Lifestyle", True),
    ("Entertainment", "Lifestyle", True),
    ("Shopping - Online", "Lifestyle", True),
    ("Shopping - Offline", "Lifestyle", True),
    ("Gifts", "Lifestyle", True),
    ("Education / Courses", "Lifestyle", False),
    ("Mobile recharge", "Lifestyle", True),
    ("Laptop / Electronics", "Lifestyle", False),
    # Travel
    ("Travel / Flight", "Travel", False),
    ("Hotels / Stay", "Travel", False),
    ("Travel - Local", "Travel", True),
    # Fixed / finance outflows
    ("EMIs", "Fixed", False),
    ("Insurance", "Fixed", False),
    ("Bank Charges", "Fixed", True),
    ("Credit Card Bill", "Fixed", False),
    # Income & inflows
    ("Salary", "Income", False),
    ("Cashback", "Income", False),
    ("Refund", "Income", False),
    ("Reimbursement", "Income", False),
    ("Interest / Dividends", "Income", False),
    ("Bonus", "Income", False),
    # Transfers / investments / adjustments
    ("FD Deposit", "Finance", False),
    ("FD Maturity", "Finance", False),
    ("Investment", "Finance", False),
    ("Transfer", "Finance", False),
    ("Reconciliation", "Finance", False),
    # Catch-all (prefer a specific category when you can)
    ("Other", "Other", True),
]

# Everyday budget cats for the dashboard bar chart (keep readable)
CHART_CATEGORIES = [
    "Groceries - Online",
    "Groceries - Physical",
    "Eating outside",
    "Food delivery",
    "Cafe / Snacks",
    "Petrol",
    "Cab / Auto",
    "Salon",
    "Subscription",
    "Electricity bill",
    "House cleaning (UC)",
    "Shopping - Online",
    "Entertainment",
    "Medical",
    "Other",
]


def style_header_row(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = thin


def money_cell(cell, formula=None, value=None, editable=False, fill=None):
    if formula is not None:
        cell.value = formula
    elif value is not None:
        cell.value = value
    cell.number_format = inr
    cell.font = money_font
    cell.border = thin
    if fill is not None:
        cell.fill = fill
    else:
        cell.fill = yellow_fill if editable else calc_fill


def _label(cell, text, bold=False):
    cell.value = text
    cell.border = thin
    if bold:
        cell.font = Font(name="Calibri", size=11, bold=True)


def _section(ws, cell_ref, text, merge_to=None):
    cell = ws[cell_ref]
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")


def _clear_sheet(ws, max_row: int = 60, max_col: int = 12) -> None:
    """Wipe values, styles, charts, and merges so a sheet can be fully rebuilt."""
    ws._charts = []
    if ws.merged_cells.ranges:
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))
    if ws.max_row and ws.max_column:
        for row in ws.iter_rows(
            min_row=1,
            max_row=max(ws.max_row, max_row),
            max_col=max(ws.max_column, max_col),
        ):
            for c in row:
                c.value = None
                c.fill = PatternFill()
                c.font = Font()
                c.border = Border()
                c.number_format = "General"


# Shared formula fragments (Configuration account rows: HDFC CC=15, ICICI CC=16)
_FORM_LIQUID = (
    "=Reconciliation!C5+Reconciliation!C6+Reconciliation!C7+Reconciliation!C8"
)
_FORM_HDFC_CC = (
    '=Configuration!C15+SUMIF(Ledger!$H:$H,"HDFC Credit Card",Ledger!$G:$G)'
    '-SUMIF(Ledger!$I:$I,"HDFC Credit Card",Ledger!$G:$G)'
)
_FORM_ICICI_CC = (
    '=Configuration!C16+SUMIF(Ledger!$H:$H,"ICICI Credit Card",Ledger!$G:$G)'
    '-SUMIF(Ledger!$I:$I,"ICICI Credit Card",Ledger!$G:$G)'
)
# Next month's budget from Monthly Budget grid, else Configuration default
_FORM_NEXT_MONTH_BUDGET = (
    "=IFERROR(INDEX('Monthly Budget'!$B$20:$B$67,"
    "MATCH(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),1),"
    "'Monthly Budget'!$A$20:$A$67,0)),Configuration!B6)"
)

# Planned Expenses summary anchors (stable cross-sheet refs for dashboards / API docs)
# Sheet layout: summary B5:B10, recurring rows 15–64, one-time rows 68–97
_PE_MONTHLY_FIXED = "='Planned Expenses'!B5"
_PE_RECURRING_COUNT = "='Planned Expenses'!B6"
_PE_YEARLY_COMMIT = "='Planned Expenses'!B7"
_PE_UPCOMING_30 = "='Planned Expenses'!B8"
_PE_UPCOMING_90 = "='Planned Expenses'!B9"
_PE_ONE_TIME_TOTAL = "='Planned Expenses'!B10"

# Sample seed rows for a fresh Planned Expenses sheet (planning only — never Ledger).
_PE_RECURRING_SEED = [
    # expense, category, frequency, amount, start, end, active, notes
    ("Rent", "Rent", "Monthly", 11000, ddate(2026, 8, 1), None, True, ""),
    ("Electricity", "Electricity bill", "Monthly", 1500, ddate(2026, 8, 1), None, True, "estimate"),
    ("Internet", "Internet / WiFi", "Monthly", 900, ddate(2026, 8, 1), None, True, ""),
    ("Spotify", "Subscription - OTT", "Monthly", 119, None, None, True, ""),
    ("Bike Insurance", "Insurance", "Yearly", 1800, ddate(2027, 1, 1), None, True, ""),
    ("Domain Renewal", "Subscription - Software", "Yearly", 3200, ddate(2027, 5, 1), None, True, ""),
]
_PE_ONETIME_SEED = [
    # expense, category, expected_month, expected_date, amount, priority, status, notes
    (
        "Shift to Hyderabad",
        "Other",
        ddate(2026, 9, 1),
        ddate(2026, 9, 15),
        25000,
        "High",
        "Planned",
        "",
    ),
    (
        "Laptop Purchase",
        "Laptop / Electronics",
        ddate(2026, 11, 1),
        None,
        120000,
        "Medium",
        "Planned",
        "",
    ),
    (
        "Vacation",
        "Travel / Flight",
        ddate(2026, 12, 1),
        None,
        35000,
        "Low",
        "Planned",
        "",
    ),
    (
        "OSCP Exam",
        "Education / Courses",
        ddate(2027, 1, 1),
        None,
        170000,
        "High",
        "Planned",
        "",
    ),
]

# Recurring Kind (cash-due buckets on dashboards). Investment is reserved.
PE_KIND_LOAN = "Loan / EMI"
PE_KIND_LIFESTYLE = "Lifestyle"
PE_KIND_INVESTMENT = "Investment"
PE_KIND_CHOICES = (PE_KIND_LOAN, PE_KIND_LIFESTYLE, PE_KIND_INVESTMENT)

# NEXT 6 MONTHS cash-due table on Planned Expenses (stable dashboard refs)
_PE_FORECAST_ORIGIN = "L4"
_PE_FORECAST_MONTH_ROW0 = 6  # L6:P11 = six months; P6 = this month total


def _infer_pe_kind(category: str | None) -> str:
    cat = (category or "").strip()
    if cat == "EMIs":
        return PE_KIND_LOAN
    if cat == "Investment":
        return PE_KIND_INVESTMENT
    return PE_KIND_LIFESTYLE


def _pe_kind_column(ws) -> int:
    """Kind lives in the recurring header row (14). Live book uses K (Payment Method is I)."""
    for c in range(1, 16):
        v = ws.cell(14, c).value
        if v and str(v).strip().lower() == "kind":
            return c
    # Template: Notes at I → Kind at J. Live: Payment Method + Notes → Kind at K.
    notes_at = None
    for c in range(1, 16):
        v = ws.cell(14, c).value
        if v and str(v).strip().lower() == "notes":
            notes_at = c
            break
    return (notes_at + 1) if notes_at else 10


def _pe_kind_match_expr(kind_col_letter: str, kind: str) -> str:
    k = f"${kind_col_letter}$15:${kind_col_letter}$64"
    b = "$B$15:$B$64"
    if kind == PE_KIND_LOAN:
        return f'(({k}="{PE_KIND_LOAN}")+(({k}="")*({b}="EMIs")))'
    if kind == PE_KIND_INVESTMENT:
        return f'(({k}="{PE_KIND_INVESTMENT}")+(({k}="")*({b}="Investment")))'
    return (
        f'(({k}="{PE_KIND_LIFESTYLE}")'
        f'+(({k}="")*({b}<>"EMIs")*({b}<>"Investment")))'
    )


def _pe_cash_due_formula(kind_col_letter: str, month_cell: str, kind: str) -> str:
    """Cash due in month_cell (1st of month) for one Kind. Respects Active + Start/End."""
    km = _pe_kind_match_expr(kind_col_letter, kind)
    m = month_cell
    monthly = (
        f"SUMPRODUCT("
        f"--($A$15:$A$64<>\"\"),"
        f"--(($G$15:$G$64=TRUE)+($G$15:$G$64=\"TRUE\")),"
        f"--({km}),"
        f"--($C$15:$C$64=\"Monthly\"),"
        f"--(($E$15:$E$64=\"\")+(($E$15:$E$64<>\"\")*($E$15:$E$64<=EOMONTH({m},0)))),"
        f"--(($F$15:$F$64=\"\")+(($F$15:$F$64<>\"\")*($F$15:$F$64>={m}))),"
        f"$D$15:$D$64)"
    )
    yearly = (
        f"SUMPRODUCT("
        f"--($A$15:$A$64<>\"\"),"
        f"--(($G$15:$G$64=TRUE)+($G$15:$G$64=\"TRUE\")),"
        f"--({km}),"
        f"--($C$15:$C$64=\"Yearly\"),"
        f"--($E$15:$E$64<>\"\"),"
        f"--(MONTH($E$15:$E$64)=MONTH({m})),"
        f"--(DATE(YEAR($E$15:$E$64),MONTH($E$15:$E$64),1)<={m}),"
        f"--(($F$15:$F$64=\"\")+(($F$15:$F$64<>\"\")*($F$15:$F$64>={m}))),"
        f"$D$15:$D$64)"
    )
    return f"={monthly}+{yearly}"


def _write_planned_forecast_table(ws, kind_col: int) -> None:
    """Write NEXT 6 MONTHS cash-due table at L4:P13. Does not touch recurring rows."""
    kind_letter = get_column_letter(kind_col)
    for ref in ("L4:P4", "L13:P13"):
        for mr in list(ws.merged_cells.ranges):
            if str(mr) == ref:
                ws.unmerge_cells(str(mr))
                break
    _section(ws, "L4", "NEXT 6 MONTHS — cash due (Active + Start/End)", "P4")
    headers = ["Month", PE_KIND_LOAN, PE_KIND_LIFESTYLE, PE_KIND_INVESTMENT, "Total"]
    for c, h in enumerate(headers, 12):
        ws.cell(5, c, h)
    style_header_row(ws, 5, 12, 16)

    ws["L6"] = "=DATE(YEAR(TODAY()),MONTH(TODAY()),1)"
    ws["L6"].number_format = "MMM-YYYY"
    ws["L6"].font = Font(bold=True)
    ws["L6"].border = thin
    ws["L6"].fill = calc_fill
    for i in range(1, 6):
        cell = ws.cell(6 + i, 12)
        cell.value = f"=EDATE(L6,{i})"
        cell.number_format = "MMM-YYYY"
        cell.border = thin
        cell.fill = calc_fill

    kinds = (PE_KIND_LOAN, PE_KIND_LIFESTYLE, PE_KIND_INVESTMENT)
    for i in range(6):
        r = 6 + i
        month_cell = f"L{r}"
        for k_i, kind in enumerate(kinds):
            money_cell(
                ws.cell(r, 13 + k_i),
                formula=_pe_cash_due_formula(kind_letter, month_cell, kind),
                fill=alert_fill if kind == PE_KIND_LOAN else soft_fill,
            )
        money_cell(ws.cell(r, 16), formula=f"=M{r}+N{r}+O{r}", fill=good_fill)
        ws.cell(r, 16).font = Font(bold=True)

    _label(ws.cell(12, 12), "6-month total", bold=True)
    for c, col in enumerate(("M", "N", "O", "P"), 13):
        fill = good_fill if col == "P" else (alert_fill if col == "M" else soft_fill)
        money_cell(ws.cell(12, c), formula=f"=SUM({col}6:{col}11)", fill=fill)
        if col in ("M", "P"):
            ws.cell(12, c).font = Font(bold=True, size=12)

    ws["L13"] = (
        "Cash due that month (not yearly÷12). Loan / EMI = must-pay. "
        "Lifestyle = everyday recurring you can cut. Investment is reserved. "
        "Blank Kind infers EMIs → Loan / EMI, else Lifestyle."
    )
    ws["L13"].font = muted_font
    ws.merge_cells("L13:P13")

    for col, w in zip(list("LMNOP"), [14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w


def _ensure_kind_column(ws) -> int:
    """Add Kind header + dropdown + yellow input cells if missing. Returns Kind col."""
    kind_col = _pe_kind_column(ws)
    header = ws.cell(14, kind_col).value
    if not header or str(header).strip().lower() != "kind":
        ws.cell(14, kind_col, "Kind")
        style_header_row(ws, 14, kind_col, kind_col)
    for r in range(15, 65):
        cell = ws.cell(r, kind_col)
        cell.fill = yellow_fill
        cell.border = thin
        if not str(cell.value or "").strip():
            cat = ws.cell(r, 2).value
            name = ws.cell(r, 1).value
            if name:
                cell.value = _infer_pe_kind(str(cat) if cat else "")
    # Dropdown (skip if already covering this column)
    already = False
    for dv in ws.data_validations.dataValidation:
        if dv.sqref and f"{get_column_letter(kind_col)}15" in str(dv.sqref):
            already = True
            break
    if not already:
        choices = ",".join(PE_KIND_CHOICES)
        dv_kind = DataValidation(type="list", formula1=f'"{choices}"', allow_blank=True)
        ws.add_data_validation(dv_kind)
        letter = get_column_letter(kind_col)
        dv_kind.add(f"{letter}15:{letter}64")
    ws.column_dimensions[get_column_letter(kind_col)].width = 14
    return kind_col


def populate_planned_expenses(ws, *, seed: bool = True) -> None:
    """Planning-only sheet: recurring + one-time expected costs. Never touches Ledger."""
    _clear_sheet(ws, max_row=120, max_col=12)
    ws._charts = []

    ws["A1"] = "PLANNED EXPENSES"
    ws["A1"].font = title_font
    ws["A2"] = (
        "Planning only — does not create Ledger entries, change balances, "
        "or affect Monthly Budget / Reconciliation. Edit yellow cells; green = formula."
    )
    ws["A2"].font = muted_font
    ws.merge_cells("A2:I2")

    # ── SUMMARY (fixed anchors B5:B10 for dashboards) ─────────────────
    _section(ws, "A4", "SUMMARY — formulas only (Active recurring + Planned one-time)", "B4")
    # Active matches Configuration style: store/display TRUE|FALSE text (also accept boolean).
    summary_rows = [
        (
            5,
            "Monthly Fixed Cost",
            # This month's recurring cash due (first month of NEXT 6 MONTHS table)
            "=P6",
        ),
        (
            6,
            "Monthly Recurring Count",
            '=COUNTIFS($G$15:$G$64,TRUE,$A$15:$A$64,"<>")'
            '+COUNTIFS($G$15:$G$64,"TRUE",$A$15:$A$64,"<>")',
        ),
        (
            7,
            "Yearly Commitments",
            '=SUMIFS($D$15:$D$64,$C$15:$C$64,"Yearly",$G$15:$G$64,TRUE)'
            '+SUMIFS($D$15:$D$64,$C$15:$C$64,"Yearly",$G$15:$G$64,"TRUE")',
        ),
        (
            8,
            "Upcoming One-Time (Next 30 Days)",
            '=SUMIFS($E$68:$E$97,$G$68:$G$97,"Planned",$I$68:$I$97,">="&TODAY(),$I$68:$I$97,"<="&TODAY()+30)',
        ),
        (
            9,
            "Upcoming One-Time (Next 90 Days)",
            '=SUMIFS($E$68:$E$97,$G$68:$G$97,"Planned",$I$68:$I$97,">="&TODAY(),$I$68:$I$97,"<="&TODAY()+90)',
        ),
        (
            10,
            "Total Planned One-Time",
            '=SUMIF($G$68:$G$97,"Planned",$E$68:$E$97)',
        ),
    ]
    for r, label, formula in summary_rows:
        _label(ws.cell(r, 1), label, bold=True)
        if "Count" in label:
            ws.cell(r, 2).value = formula
            ws.cell(r, 2).font = big_num_font
            ws.cell(r, 2).fill = soft_fill
            ws.cell(r, 2).border = thin
            ws.cell(r, 2).number_format = "0"
        else:
            money_cell(ws.cell(r, 2), formula=formula, fill=soft_fill)
            ws.cell(r, 2).font = big_num_font

    ws["A11"] = (
        "Monthly Fixed Cost = this month's recurring cash due (P6): Active rows whose "
        "Start/End cover this month. Yearly amounts hit their due month, not yearly÷12. "
        "Kind splits Loan / EMI vs Lifestyle vs Investment. One-time windows use Effective Date."
    )
    ws["A11"].font = muted_font
    ws.merge_cells("A11:K11")

    # ── SECTION 1 — RECURRING ─────────────────────────────────────────
    _section(ws, "A13", "RECURRING EXPENSES", "J13")
    rec_headers = [
        "Expense",
        "Category",
        "Frequency",
        "Amount",
        "Start",
        "End",
        "Active",
        "Monthly Equivalent",
        "Notes",
        "Kind",
    ]
    for c, h in enumerate(rec_headers, 1):
        ws.cell(14, c, h)
    style_header_row(ws, 14, 1, 10)

    REC_START, REC_END = 15, 64
    for r in range(REC_START, REC_END + 1):
        for c in (1, 2, 3, 5, 6, 7, 9, 10):
            ws.cell(r, c).fill = yellow_fill
            ws.cell(r, c).border = thin
        money_cell(ws.cell(r, 4), value=None, editable=True)
        # Monthly Equivalent: always from Frequency/Amount (Active + dates filter cash-due)
        money_cell(
            ws.cell(r, 8),
            formula=(
                f'=IF(D{r}="","",'
                f'IF(C{r}="Monthly",D{r},'
                f'IF(C{r}="Yearly",ROUND(D{r}/12,2),"")))'
            ),
        )
        ws.cell(r, 5).number_format = "MMM-YYYY"
        ws.cell(r, 6).number_format = "MMM-YYYY"

    if seed:
        for i, row in enumerate(_PE_RECURRING_SEED):
            r = REC_START + i
            exp, cat, freq, amt, start, end, active, notes = row
            ws.cell(r, 1, exp)
            ws.cell(r, 2, cat)
            ws.cell(r, 3, freq)
            money_cell(ws.cell(r, 4), value=amt, editable=True)
            if start is not None:
                ws.cell(r, 5, start)
            if end is not None:
                ws.cell(r, 6, end)
            ws.cell(r, 7, "TRUE" if active else "FALSE")
            ws.cell(r, 9, notes)
            ws.cell(r, 10, _infer_pe_kind(cat))

    # ── SECTION 2 — ONE-TIME ──────────────────────────────────────────
    _section(ws, "A66", "UPCOMING ONE-TIME EXPENSES", "I66")
    ot_headers = [
        "Expense",
        "Category",
        "Expected Month",
        "Expected Date",
        "Amount",
        "Priority",
        "Status",
        "Notes",
        "Effective Date",
    ]
    for c, h in enumerate(ot_headers, 1):
        ws.cell(67, c, h)
    style_header_row(ws, 67, 1, 9)

    OT_START, OT_END = 68, 97
    for r in range(OT_START, OT_END + 1):
        for c in (1, 2, 3, 4, 6, 7, 8):
            ws.cell(r, c).fill = yellow_fill
            ws.cell(r, c).border = thin
        money_cell(ws.cell(r, 5), value=None, editable=True)
        # Effective Date for 30/90 SUMIFS (exact date preferred, else month)
        ws.cell(r, 9).value = (
            f'=IF(A{r}="","",IF(D{r}<>"",D{r},IF(C{r}<>"",C{r},"")))'
        )
        ws.cell(r, 9).fill = calc_fill
        ws.cell(r, 9).border = thin
        ws.cell(r, 9).number_format = "dd-mmm-yyyy"
        ws.cell(r, 3).number_format = "MMM-YYYY"
        ws.cell(r, 4).number_format = "dd-mmm-yyyy"

    if seed:
        for i, row in enumerate(_PE_ONETIME_SEED):
            r = OT_START + i
            exp, cat, emonth, edate, amt, prio, status, notes = row
            ws.cell(r, 1, exp)
            ws.cell(r, 2, cat)
            if emonth is not None:
                ws.cell(r, 3, emonth)
            if edate is not None:
                ws.cell(r, 4, edate)
            money_cell(ws.cell(r, 5), value=amt, editable=True)
            ws.cell(r, 6, prio)
            ws.cell(r, 7, status)
            ws.cell(r, 8, notes)

    # Data validation (lists)
    # Clear any prior validations when rebuilding this sheet alone
    ws.data_validations.dataValidation = []
    dv_freq = DataValidation(type="list", formula1='"Monthly,Yearly"', allow_blank=True)
    dv_active = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_prio = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv_status = DataValidation(
        type="list", formula1='"Planned,Completed,Cancelled"', allow_blank=True
    )
    dv_kind = DataValidation(
        type="list",
        formula1=f'"{",".join(PE_KIND_CHOICES)}"',
        allow_blank=True,
    )
    # Categories from Configuration (same list as Ledger)
    cat_start = 27
    cat_end = 27 + len(CATEGORIES) + 15 - 1  # matches patch_live blank slots
    dv_cat = DataValidation(
        type="list",
        formula1=f"Configuration!$A${cat_start}:$A${cat_end}",
        allow_blank=True,
    )
    for dv in (dv_freq, dv_active, dv_prio, dv_status, dv_kind, dv_cat):
        ws.add_data_validation(dv)
    dv_freq.add(f"C{REC_START}:C{REC_END}")
    dv_active.add(f"G{REC_START}:G{REC_END}")
    dv_kind.add(f"J{REC_START}:J{REC_END}")
    dv_prio.add(f"F{OT_START}:F{OT_END}")
    dv_status.add(f"G{OT_START}:G{OT_END}")
    dv_cat.add(f"B{REC_START}:B{REC_END}")
    dv_cat.add(f"B{OT_START}:B{OT_END}")

    _write_planned_forecast_table(ws, kind_col=10)

    ws["A99"] = (
        "Tips: Kind = Loan / EMI (must-pay) vs Lifestyle (negotiable) vs Investment. "
        "Set Active=FALSE to drop a row from cash-due. End date stops EMIs after the last month. "
        "Mark one-time Status=Completed/Cancelled when done. This sheet never posts to the Ledger."
    )
    ws["A99"].font = muted_font
    ws.merge_cells("A99:J99")

    widths = {
        "A": 22,
        "B": 22,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 18,
        "I": 14,
        "J": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A15"
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[67].height = 30


def _add_planned_summary_block(dash, start_row: int, *, cols: str = "D") -> int:
    """
    Write Planned Expenses summary block starting at start_row.
    Returns the last row written. Surgical — does not clear other content.
    cols: merge end column letter for section header (e.g. 'D' or 'F').
    """
    end = cols
    _section(
        dash,
        f"A{start_row}",
        "PLANNED EXPENSES — planning only (not in Ledger / budget)",
        f"{end}{start_row}",
    )
    r = start_row + 1
    rows = [
        ("Monthly fixed cost", _PE_MONTHLY_FIXED, "this month cash due (Start/End + Kind)"),
        ("Monthly recurring count", _PE_RECURRING_COUNT, "active rows"),
        ("Yearly commitments", _PE_YEARLY_COMMIT, "sum of active yearly amounts"),
        ("Upcoming one-time (30 days)", _PE_UPCOMING_30, "Status=Planned"),
        ("Upcoming one-time (90 days)", _PE_UPCOMING_90, "Status=Planned"),
        ("Total planned one-time", _PE_ONE_TIME_TOTAL, "all Status=Planned"),
    ]
    for label, formula, note in rows:
        _label(dash.cell(r, 1), label, bold=("fixed" in label.lower() or "total" in label.lower()))
        if "count" in label.lower():
            dash.cell(r, 2).value = formula
            dash.cell(r, 2).font = big_num_font
            dash.cell(r, 2).fill = soft_fill
            dash.cell(r, 2).border = thin
            dash.cell(r, 2).number_format = "0"
        else:
            fill = good_fill if "fixed" in label.lower() else soft_fill
            if "30" in label or "90" in label:
                fill = alert_fill if "30" in label else soft_fill
            money_cell(dash.cell(r, 2), formula=formula, fill=fill)
            if "fixed" in label.lower() or "total" in label.lower():
                dash.cell(r, 2).font = Font(bold=True, size=12)
        dash.cell(r, 3).value = note
        dash.cell(r, 3).font = muted_font
        r += 1
    dash.cell(r, 1).value = (
        "Edit lists on the Planned Expenses sheet. Active=FALSE excludes a recurring row "
        "from monthly fixed cost. Does not create Ledger entries."
    )
    dash.cell(r, 1).font = muted_font
    dash.merge_cells(f"A{r}:{end}{r}")
    return r


def _add_commitment_forecast_block(dash, start_row: int, *, cols: str = "E") -> int:
    """
    NEXT 6 MONTHS recurring cash-due table. Refs Planned Expenses!L6:P12.
    Surgical — does not clear other content. Returns last row written.
    """
    end = cols
    _section(
        dash,
        f"A{start_row}",
        "NEXT 6 MONTHS — recurring cash due",
        f"{end}{start_row}",
    )
    hr = start_row + 1
    headers = ["Month", PE_KIND_LOAN, PE_KIND_LIFESTYLE, PE_KIND_INVESTMENT, "Total"]
    for c, h in enumerate(headers, 1):
        dash.cell(hr, c, h)
    style_header_row(dash, hr, 1, 5)

    for i in range(6):
        r = hr + 1 + i
        src = _PE_FORECAST_MONTH_ROW0 + i
        dash.cell(r, 1).value = f"='Planned Expenses'!L{src}"
        dash.cell(r, 1).number_format = "MMM-YYYY"
        dash.cell(r, 1).border = thin
        dash.cell(r, 1).fill = calc_fill
        money_cell(
            dash.cell(r, 2),
            formula=f"='Planned Expenses'!M{src}",
            fill=alert_fill,
        )
        money_cell(
            dash.cell(r, 3),
            formula=f"='Planned Expenses'!N{src}",
            fill=soft_fill,
        )
        money_cell(
            dash.cell(r, 4),
            formula=f"='Planned Expenses'!O{src}",
            fill=soft_fill,
        )
        money_cell(
            dash.cell(r, 5),
            formula=f"='Planned Expenses'!P{src}",
            fill=good_fill,
        )
        dash.cell(r, 5).font = Font(bold=True)

    tot = hr + 7
    _label(dash.cell(tot, 1), "6-month total", bold=True)
    money_cell(dash.cell(tot, 2), formula="='Planned Expenses'!M12", fill=alert_fill)
    money_cell(dash.cell(tot, 3), formula="='Planned Expenses'!N12", fill=soft_fill)
    money_cell(dash.cell(tot, 4), formula="='Planned Expenses'!O12", fill=soft_fill)
    money_cell(dash.cell(tot, 5), formula="='Planned Expenses'!P12", fill=good_fill)
    dash.cell(tot, 2).font = Font(bold=True)
    dash.cell(tot, 5).font = Font(bold=True, size=12)

    note = tot + 1
    dash.cell(note, 1).value = (
        "Loan / EMI = must-pay (EMI, loans). Lifestyle = everyday recurring you can cut. "
        "Investment is reserved (0 until you tag rows). Dates from Planned Expenses Start/End. "
        "MacBook SmartEMI is Loan / EMI for Sep-2026 through Feb-2027."
    )
    dash.cell(note, 1).font = muted_font
    dash.merge_cells(f"A{note}:{end}{note}")
    dash.row_dimensions[note].height = 32
    for col_letter, w in zip(list("ABCDE"), [22, 14, 14, 14, 14]):
        current = dash.column_dimensions[col_letter].width
        if not current or current < w:
            dash.column_dimensions[col_letter].width = w
    return note


def _dashboard_has_block(dash, needle: str, max_row: int = 120) -> bool:
    key = needle.upper()
    for r in range(1, min(dash.max_row or 1, max_row) + 1):
        v = dash.cell(r, 1).value
        if v and key in str(v).upper():
            return True
    return False


def populate_simple_dashboard(dash) -> None:
    """Minimal dashboard: month pace, CC bills, free-to-allocate, next-month estimate."""
    _clear_sheet(dash, max_row=50, max_col=6)

    dash["A1"] = "SIMPLE DASHBOARD"
    dash["A1"].font = title_font
    dash["A2"] = (
        "Only what matters day-to-day. Charts & full balances live on Detailed Dashboard."
    )
    dash["A2"].font = muted_font
    dash.merge_cells("A2:D2")

    # ── THIS MONTH / PACE ──────────────────────────────────────────────
    _section(dash, "A4", "THIS MONTH — PACE", "D4")

    _label(dash["A5"], "Month")
    dash["B5"] = '=TEXT(TODAY(),"MMMM YYYY")'
    dash["B5"].font = Font(bold=True, size=12, color="1F4E79")
    dash["B5"].border = thin

    _label(dash["C5"], "Days left")
    # Include today; integer format (LibreOffice otherwise shows a date serial).
    dash["D5"] = "=EOMONTH(TODAY(),0)-TODAY()+1"
    dash["D5"].font = big_num_font
    dash["D5"].fill = soft_fill
    dash["D5"].border = thin
    dash["D5"].number_format = "0"
    dash["D5"].alignment = Alignment(horizontal="center")

    _label(dash["A6"], "Budget remaining", bold=True)
    money_cell(dash["B6"], formula="='Monthly Budget'!B9", fill=good_fill)
    dash["B6"].font = big_num_font

    _label(dash["C6"], "Safe to spend / day", bold=True)
    money_cell(
        dash["D6"],
        formula="=IF(D5<=0,0,MAX(0,B6)/D5)",
        fill=soft_fill,
    )
    dash["D6"].font = big_num_font

    _label(dash["A7"], "Budget used %")
    dash["B7"] = "='Monthly Budget'!B16"
    dash["B7"].number_format = pct
    dash["B7"].fill = calc_fill
    dash["B7"].border = thin

    _label(dash["C7"], "Month elapsed %")
    dash["D7"] = "=DAY(TODAY())/DAY(EOMONTH(TODAY(),0))"
    dash["D7"].number_format = pct
    dash["D7"].fill = calc_fill
    dash["D7"].border = thin

    _label(dash["A8"], "Pace check", bold=True)
    dash["B8"] = (
        '=IF(B6<0,"OVER BUDGET — stop discretionary spend",'
        'IF(B7>D7+0.05,"Spending faster than the month — slow down",'
        'IF(B7>D7,"Slightly ahead of pace — be careful","On track")))'
    )
    dash["B8"].font = warn_font
    dash["B8"].fill = alert_fill
    dash["B8"].border = thin
    dash.merge_cells("B8:D8")

    dash["A9"] = (
        "Tip: Safe ₹/day = budget remaining ÷ days left (incl. today). "
        "If used % > month elapsed %, you are burning budget too fast."
    )
    dash["A9"].font = muted_font
    dash.merge_cells("A9:D9")

    # ── UPCOMING CREDIT CARD BILLS ─────────────────────────────────────
    _section(dash, "A11", "UPCOMING CREDIT CARD BILLS", "D11")

    _label(dash["A12"], "HDFC Credit Card")
    money_cell(dash["B12"], formula=_FORM_HDFC_CC)
    dash["C12"] = "outstanding due"
    dash["C12"].font = muted_font

    _label(dash["A13"], "ICICI Credit Card")
    money_cell(dash["B13"], formula=_FORM_ICICI_CC)
    dash["C13"] = "outstanding due"
    dash["C13"].font = muted_font

    _label(dash["A14"], "Total CC due", bold=True)
    money_cell(dash["B14"], formula="=B12+B13", fill=alert_fill)
    dash["B14"].font = Font(bold=True, size=12)

    dash["A15"] = "Pay from savings before treating anything as free to allocate."
    dash["A15"].font = muted_font
    dash.merge_cells("A15:D15")

    # ── FREE TO ALLOCATE (savings − budget still reserved) ─────────────
    _section(dash, "A17", "FREE TO ALLOCATE — emergency · invest · goals", "D17")

    _label(dash["A18"], "Total liquid savings")
    # HDFC + ICICI + Cash + Wallet (Reconciliation rows match ACCOUNTS order)
    money_cell(dash["B18"], formula=_FORM_LIQUID)
    dash["C18"] = "savings + cash + wallets"
    dash["C18"].font = muted_font

    _label(dash["A19"], "Budget still reserved")
    money_cell(dash["B19"], formula="=MAX(0,B6)")
    dash["C19"] = "money still needed for this month's budget"
    dash["C19"].font = muted_font

    _label(dash["A20"], "Free to allocate", bold=True)
    money_cell(dash["B20"], formula="=B18-B19", fill=good_fill)
    dash["B20"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    dash["C20"] = "total savings − budget remaining"
    dash["C20"].font = muted_font

    dash["A21"] = (
        "This is leftover cash after earmarking what the month still needs. "
        "Use it for Emergency fund, Investments, or Goals (details coming next). "
        "It does not subtract CC dues — pay those first if due soon."
    )
    dash["A21"].font = muted_font
    dash.merge_cells("A21:D21")
    dash.row_dimensions[21].height = 36

    # ── NEXT MONTH FREE-TO-ALLOCATE (estimate) ─────────────────────────
    _section(dash, "A23", "NEXT MONTH FREE-TO-ALLOCATE — estimate", "D23")

    _label(dash["A24"], "Free to allocate (today)")
    money_cell(dash["B24"], formula="=B20")
    dash["C24"] = "starting point"
    dash["C24"].font = muted_font

    _label(dash["A25"], "− Pay all CC dues")
    money_cell(dash["B25"], formula="=B14")
    dash["C25"] = "after clearing upcoming bills"
    dash["C25"].font = muted_font

    _label(dash["A26"], "+ Monthly salary")
    money_cell(dash["B26"], formula="=Configuration!B7")
    dash["C26"] = "from Configuration (edit yellow cell)"
    dash["C26"].font = muted_font

    _label(dash["A27"], "− Next month budget")
    money_cell(dash["B27"], formula=_FORM_NEXT_MONTH_BUDGET)
    dash["C27"] = "reserved for next month's spend cap"
    dash["C27"].font = muted_font

    _label(dash["A28"], "Est. free next month", bold=True)
    money_cell(dash["B28"], formula="=B24-B25+B26-B27", fill=soft_fill)
    dash["B28"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    dash["C28"] = "potential surplus after CC + budget + salary"
    dash["C28"].font = muted_font

    dash["A29"] = (
        "Estimate = free-to-allocate today − CC dues + monthly salary − next month's budget. "
        "Assumes you finish this month's remaining budget and pay full CC dues. "
        "If this month's salary is already in liquid savings, it is already counted — "
        "use Configuration monthly salary as take-home you expect to receive for the next cycle."
    )
    dash["A29"].font = muted_font
    dash.merge_cells("A29:D29")
    dash.row_dimensions[29].height = 48

    # ── PLANNED EXPENSES (planning only) ───────────────────────────────
    pe_end = _add_planned_summary_block(dash, 31, cols="D")
    _add_commitment_forecast_block(dash, pe_end + 2, cols="E")

    for col_letter, w in zip(list("ABCD"), [28, 16, 40, 14]):
        dash.column_dimensions[col_letter].width = w
    dash.row_dimensions[8].height = 22
    dash.freeze_panes = "A4"


def populate_detailed_dashboard(dash, n_accounts: int | None = None) -> None:
    """Fill Detailed Dashboard: pace cards, balances, net worth, 3 charts."""
    if n_accounts is None:
        n_accounts = len(ACCOUNTS)

    _clear_sheet(dash, max_row=60, max_col=12)

    dash["A1"] = "DETAILED DASHBOARD"
    dash["A1"].font = title_font
    dash["A2"] = (
        "Full snapshot + charts. Day-to-day numbers live on Simple Dashboard. "
        "Reload after phone saves."
    )
    dash["A2"].font = muted_font
    dash.merge_cells("A2:F2")

    # ── PACE / BE CAREFUL ──────────────────────────────────────────────
    _section(dash, "A4", "PACE — BE CAREFUL", "F4")

    _label(dash["A5"], "Month")
    dash["B5"] = '=TEXT(TODAY(),"MMMM YYYY")'
    dash["B5"].font = Font(bold=True, size=12, color="1F4E79")
    dash["B5"].border = thin

    _label(dash["C5"], "Days left in month")
    # Include today in remaining so "1 day left" on last day still shows budget/day.
    # Force integer format — LibreOffice otherwise shows the day-count as a date (e.g. 25/01/00).
    dash["D5"] = "=EOMONTH(TODAY(),0)-TODAY()+1"
    dash["D5"].font = big_num_font
    dash["D5"].fill = soft_fill
    dash["D5"].border = thin
    dash["D5"].number_format = "0"
    dash["D5"].alignment = Alignment(horizontal="center")

    _label(dash["E5"], "Days elapsed")
    dash["F5"] = "=DAY(TODAY())"
    dash["F5"].fill = calc_fill
    dash["F5"].border = thin
    dash["F5"].number_format = "0"
    dash["F5"].alignment = Alignment(horizontal="center")

    _label(dash["A6"], "Monthly Budget")
    money_cell(dash["B6"], formula="='Monthly Budget'!B6")

    _label(dash["C6"], "Budget remaining", bold=True)
    money_cell(dash["D6"], formula="='Monthly Budget'!B9", fill=good_fill)
    dash["D6"].font = big_num_font

    _label(dash["E6"], "Budget spent")
    money_cell(dash["F6"], formula="='Monthly Budget'!B8")

    _label(dash["A7"], "Safe to spend / day left", bold=True)
    money_cell(
        dash["B7"],
        formula='=IF(D5<=0,0,MAX(0,D6)/D5)',
        fill=soft_fill,
    )
    dash["B7"].font = big_num_font

    _label(dash["C7"], "Budget used %")
    dash["D7"] = "='Monthly Budget'!B16"
    dash["D7"].number_format = pct
    dash["D7"].fill = calc_fill
    dash["D7"].border = thin

    _label(dash["E7"], "Month elapsed %")
    dash["F7"] = "=DAY(TODAY())/DAY(EOMONTH(TODAY(),0))"
    dash["F7"].number_format = pct
    dash["F7"].fill = calc_fill
    dash["F7"].border = thin

    _label(dash["A8"], "Pace check", bold=True)
    dash["B8"] = (
        '=IF(D6<0,"OVER BUDGET — stop discretionary spend",'
        'IF(D7>F7+0.05,"Spending faster than the month — slow down",'
        'IF(D7>F7,"Slightly ahead of pace — be careful","On track")))'
    )
    dash["B8"].font = warn_font
    dash["B8"].fill = alert_fill
    dash["B8"].border = thin
    dash.merge_cells("B8:F8")

    dash["A9"] = (
        "Tip: Safe ₹/day = budget remaining ÷ days left (including today). "
        "If used % is above month elapsed %, you are burning budget too fast."
    )
    dash["A9"].font = muted_font
    dash.merge_cells("A9:F9")

    # ── THIS MONTH snapshot ────────────────────────────────────────────
    _section(dash, "A11", "THIS MONTH", "B11")
    for label, formula, r, is_pct in [
        ("Income", "='Monthly Budget'!B7", 12, False),
        ("Budget Expenses", "='Monthly Budget'!B8", 13, False),
        ("Non-Budget Expenses", "='Monthly Budget'!B10", 14, False),
        ("Est. Savings", "='Monthly Budget'!B15", 15, False),
        ("Investments", "='Monthly Budget'!B11", 16, False),
        ("EMIs + Rent", "='Monthly Budget'!B12+'Monthly Budget'!B13", 17, False),
    ]:
        _label(dash.cell(r, 1), label)
        if is_pct:
            dash.cell(r, 2, formula)
            dash.cell(r, 2).number_format = pct
            dash.cell(r, 2).fill = calc_fill
            dash.cell(r, 2).border = thin
        else:
            money_cell(dash.cell(r, 2), formula=formula)

    # ── CREDIT CARDS (upcoming bills) ──────────────────────────────────
    _section(dash, "D11", "UPCOMING CREDIT CARD BILLS", "F11")
    dash["D12"] = "HDFC Credit Card"
    dash["D12"].font = Font(bold=True)
    _label(dash["D13"], "Outstanding")
    money_cell(dash["E13"], formula=_FORM_HDFC_CC)
    _label(dash["D14"], "Limit / Available")
    money_cell(dash["E14"], formula="=Configuration!D15")
    money_cell(dash["F14"], formula="=E14-E13")
    _label(dash["D15"], "Utilization")
    dash["E15"] = "=IF(E14=0,0,E13/E14)"
    dash["E15"].number_format = pct
    dash["E15"].fill = calc_fill
    dash["E15"].border = thin

    dash["D16"] = "ICICI Credit Card"
    dash["D16"].font = Font(bold=True)
    _label(dash["D17"], "Outstanding")
    money_cell(dash["E17"], formula=_FORM_ICICI_CC)
    _label(dash["D18"], "Limit / Available")
    money_cell(dash["E18"], formula="=Configuration!D16")
    money_cell(dash["F18"], formula="=E18-E17")
    _label(dash["D19"], "Utilization")
    dash["E19"] = "=IF(E18=0,0,E17/E18)"
    dash["E19"].number_format = pct
    dash["E19"].fill = calc_fill
    dash["E19"].border = thin

    _label(dash["D20"], "Total CC due", bold=True)
    money_cell(dash["E20"], formula="=E13+E17", fill=alert_fill)
    dash["E20"].font = Font(bold=True, size=12)

    # ── FREE TO ALLOCATE ───────────────────────────────────────────────
    _section(dash, "A19", "FREE TO ALLOCATE — emergency · invest · goals", "B19")
    _label(dash["A20"], "Total liquid savings")
    money_cell(dash["B20"], formula=_FORM_LIQUID)
    _label(dash["A21"], "Budget still reserved")
    money_cell(dash["B21"], formula="=MAX(0,D6)")
    _label(dash["A22"], "Free to allocate", bold=True)
    money_cell(dash["B22"], formula="=B20-B21", fill=good_fill)
    dash["B22"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    # ── NEXT MONTH FREE-TO-ALLOCATE (estimate) ─────────────────────────
    _section(dash, "D22", "NEXT MONTH FREE-TO-ALLOCATE — estimate", "F22")
    _label(dash["D23"], "Free to allocate (today)")
    money_cell(dash["E23"], formula="=B22")
    _label(dash["D24"], "− Pay all CC dues")
    money_cell(dash["E24"], formula="=E20")
    _label(dash["D25"], "+ Monthly salary")
    money_cell(dash["E25"], formula="=Configuration!B7")
    _label(dash["D26"], "− Next month budget")
    money_cell(dash["E26"], formula=_FORM_NEXT_MONTH_BUDGET)
    _label(dash["D27"], "Est. free next month", bold=True)
    money_cell(dash["E27"], formula="=E23-E24+E25-E26", fill=soft_fill)
    dash["E27"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    dash["A23"] = (
        "Free to allocate = liquid − budget remaining (does not subtract CC). "
        "Next-month estimate = free today − CC dues + salary − next budget. "
        "Set Monthly Salary on Configuration."
    )
    dash["A23"].font = muted_font
    dash.merge_cells("A23:B27")
    dash.row_dimensions[23].height = 48

    # ── ACCOUNT BALANCES + NET WORTH ───────────────────────────────────
    bal_section = 29
    _section(dash, f"A{bal_section}", "ACCOUNT BALANCES", f"B{bal_section}")
    dash.cell(bal_section + 1, 1, "Account")
    dash.cell(bal_section + 1, 2, "Balance / Due")
    style_header_row(dash, bal_section + 1, 1, 2)
    bal_start = bal_section + 2
    for i in range(n_accounts):
        r = bal_start + i
        rec_r = 5 + i
        dash.cell(r, 1, f"=Reconciliation!A{rec_r}").border = thin
        money_cell(dash.cell(r, 2), formula=f"=Reconciliation!C{rec_r}")
    bal_end = bal_start + n_accounts - 1

    _section(dash, f"D{bal_section}", "NET WORTH", f"F{bal_section}")
    # Reconciliation rows: 5.. match ACCOUNTS order in template
    _label(dash.cell(bal_section + 1, 4), "Savings & Cash")
    money_cell(dash.cell(bal_section + 1, 5), formula=_FORM_LIQUID)
    _label(dash.cell(bal_section + 2, 4), "FD")
    money_cell(dash.cell(bal_section + 2, 5), formula="=Reconciliation!C11")
    _label(dash.cell(bal_section + 3, 4), "Investments (MF)")
    money_cell(dash.cell(bal_section + 3, 5), formula="=Reconciliation!C12")
    _label(dash.cell(bal_section + 4, 4), "Total Assets")
    money_cell(
        dash.cell(bal_section + 4, 5),
        formula=f"=E{bal_section + 1}+E{bal_section + 2}+E{bal_section + 3}",
    )
    dash.cell(bal_section + 4, 5).font = Font(bold=True)
    _label(dash.cell(bal_section + 5, 4), "Credit Card Due")
    money_cell(
        dash.cell(bal_section + 5, 5),
        formula="=Reconciliation!C9+Reconciliation!C10",
    )
    _label(dash.cell(bal_section + 6, 4), "NET WORTH", bold=True)
    money_cell(
        dash.cell(bal_section + 6, 5),
        formula=f"=E{bal_section + 4}-E{bal_section + 5}",
        fill=good_fill,
    )
    dash.cell(bal_section + 6, 5).font = Font(bold=True, size=14, color="1F4E79")

    # ── CHART DATA (right side, narrow — still visible for transparency) ─
    # Budget burn doughnut source
    dash["H4"] = "CHART DATA"
    dash["H4"].font = section_font
    dash["H4"].fill = section_fill
    dash.merge_cells("H4:I4")
    dash["H5"] = "Budget slice"
    dash["I5"] = "Amount"
    style_header_row(dash, 5, 8, 9)
    dash["H6"] = "Spent"
    money_cell(dash["I6"], formula="=MAX(0,F6)")
    dash["H7"] = "Remaining"
    money_cell(dash["I7"], formula="=MAX(0,D6)")
    dash["H8"] = "(Over budget ignored in pie; see pace check)"
    dash["H8"].font = muted_font
    dash.merge_cells("H8:I8")

    # Last 6 months income vs budget expenses
    dash["H10"] = "Last 6 months"
    dash["H10"].font = section_font
    dash["H10"].fill = section_fill
    dash.merge_cells("H10:J10")
    dash["H11"] = "Month"
    dash["I11"] = "Income"
    dash["J11"] = "Budget Exp"
    style_header_row(dash, 11, 8, 10)
    for i in range(6):
        r = 12 + i
        # i=0 → 5 months ago … i=5 → current month
        offset = i - 5
        dash.cell(
            r,
            8,
            f'=TEXT(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),{offset}),"mmm yy")',
        ).border = thin
        money_cell(
            dash.cell(r, 9),
            formula=(
                f'=SUMIFS(Ledger!$G:$G,Ledger!$D:$D,'
                f'TEXT(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),{offset}),"MMMM"),'
                f'Ledger!$E:$E,YEAR(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),{offset})),'
                f'Ledger!$F:$F,"Income")'
            ),
        )
        money_cell(
            dash.cell(r, 10),
            formula=(
                f'=SUMIFS(Ledger!$G:$G,Ledger!$D:$D,'
                f'TEXT(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),{offset}),"MMMM"),'
                f'Ledger!$E:$E,YEAR(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),{offset})),'
                f'Ledger!$F:$F,"Expense",Ledger!$K:$K,TRUE())'
                f'-SUMIFS(Ledger!$G:$G,Ledger!$D:$D,'
                f'TEXT(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),{offset}),"MMMM"),'
                f'Ledger!$E:$E,YEAR(EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),{offset})),'
                f'Ledger!$F:$F,"Refund",Ledger!$K:$K,TRUE())'
            ),
        )

    # Category spend this month (budget expenses by category)
    dash["H19"] = "Spend by category (this month)"
    dash["H19"].font = section_font
    dash["H19"].fill = section_fill
    dash.merge_cells("H19:I19")
    dash["H20"] = "Category"
    dash["I20"] = "Spent"
    style_header_row(dash, 20, 8, 9)
    chart_cats = list(CHART_CATEGORIES)
    for i, cat in enumerate(chart_cats):
        r = 21 + i
        dash.cell(r, 8, cat).border = thin
        money_cell(
            dash.cell(r, 9),
            formula=(
                f'=SUMIFS(Ledger!$G:$G,Ledger!$D:$D,TEXT(TODAY(),"MMMM"),'
                f'Ledger!$E:$E,YEAR(TODAY()),Ledger!$F:$F,"Expense",'
                f'Ledger!$J:$J,H{r})'
                f'-SUMIFS(Ledger!$G:$G,Ledger!$D:$D,TEXT(TODAY(),"MMMM"),'
                f'Ledger!$E:$E,YEAR(TODAY()),Ledger!$F:$F,"Refund",'
                f'Ledger!$J:$J,H{r})'
            ),
        )
    cat_end = 20 + len(chart_cats)

    # ── CHARTS ─────────────────────────────────────────────────────────
    # 1) Budget burn doughnut
    pie = DoughnutChart()
    pie.title = "Budget: spent vs remaining"
    labels = Reference(dash, min_col=8, min_row=6, max_row=7)
    data = Reference(dash, min_col=9, min_row=5, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = True
    pie.style = 10
    pie.width = 12
    pie.height = 8
    # series colors: spent=coral, remaining=green
    try:
        s = pie.series[0]
        pt0 = DataPoint(idx=0)
        pt0.graphicalProperties = GraphicalProperties(
            solidFill="E57373"
        )
        pt1 = DataPoint(idx=1)
        pt1.graphicalProperties = GraphicalProperties(
            solidFill="66BB6A"
        )
        s.data_points = [pt0, pt1]
    except Exception:
        pass
    dash.add_chart(pie, "A35")

    # 2) 6-month income vs expenses
    col = BarChart()
    col.type = "col"
    col.grouping = "clustered"
    col.title = "Last 6 months: income vs budget spend"
    col.y_axis.title = "₹"
    col.x_axis.title = None
    data2 = Reference(dash, min_col=9, min_row=11, max_col=10, max_row=17)
    cats2 = Reference(dash, min_col=8, min_row=12, max_row=17)
    col.add_data(data2, titles_from_data=True)
    col.set_categories(cats2)
    col.shape = 4
    col.style = 10
    col.width = 15
    col.height = 9
    dash.add_chart(col, "D35")

    # 3) Category bars
    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "This month — where budget spend went"
    bar.y_axis.title = None
    data3 = Reference(dash, min_col=9, min_row=20, max_row=cat_end)
    cats3 = Reference(dash, min_col=8, min_row=21, max_row=cat_end)
    bar.add_data(data3, titles_from_data=True)
    bar.set_categories(cats3)
    bar.shape = 4
    bar.width = 15
    bar.height = 10
    dash.add_chart(bar, "A52")

    # Quick legend under charts area notes
    note_row = bal_end + 2
    if note_row < 34:
        note_row = 34
    dash.cell(note_row, 1, "Charts sit below · Chart source tables in columns H–J")
    dash.cell(note_row, 1).font = muted_font

    # Planned Expenses summary (below balances/charts note — planning only)
    pe_row = max(note_row + 2, 45)
    pe_end = _add_planned_summary_block(dash, pe_row, cols="F")
    # Category chart sits around rows 55–73; keep the 6-month table below it.
    _add_commitment_forecast_block(dash, max(pe_end + 2, 75), cols="E")

    for col_letter, w in zip(
        list("ABCDEFGHIJ"), [26, 16, 22, 18, 16, 14, 3, 22, 12, 12]
    ):
        dash.column_dimensions[col_letter].width = w
    dash.row_dimensions[8].height = 22
    dash.freeze_panes = "A4"


def build() -> Path:
    wb = Workbook()
    cfg = wb.active
    cfg.title = "Configuration"

    # ── Configuration ──────────────────────────────────────────────────
    cfg["A1"] = "CONFIGURATION"
    cfg["A1"].font = title_font
    cfg["A2"] = (
        "Edit only yellow cells. Everything else is derived from the Ledger. "
        "Do not store current balances here — only opening balances and credit limits."
    )
    cfg["A2"].font = muted_font
    cfg.merge_cells("A2:G2")

    cfg["A4"] = "BUDGET DEFAULT"
    cfg["A4"].font = section_font
    cfg["A4"].fill = section_fill
    cfg.merge_cells("A4:C4")
    cfg["A5"] = "Setting"
    cfg["B5"] = "Value"
    cfg["C5"] = "Notes"
    style_header_row(cfg, 5, 1, 3)
    cfg["A6"] = "Default Monthly Budget (₹)"
    money_cell(cfg["B6"], value=10000, editable=True)
    cfg["C6"] = "Override per month on Monthly Budget sheet"
    cfg["A7"] = "Monthly Salary (₹)"
    money_cell(cfg["B7"], value=0, editable=True)
    cfg["C7"] = "Take-home used for next-month free-to-allocate estimate on dashboards"

    cfg["A8"] = "ACCOUNTS"
    cfg["A8"].font = section_font
    cfg["A8"].fill = section_fill
    cfg.merge_cells("A8:F8")
    cfg["A9"] = (
        "Every ledger entry moves money From Account → To Account. "
        "Opening Balance is the starting point only."
    )
    cfg["A9"].font = muted_font
    cfg.merge_cells("A9:F9")

    for i, h in enumerate(
        ["Account", "Type", "Opening Balance (₹)", "Credit Limit (₹)", "Include in Net Worth", "Notes"],
        1,
    ):
        cfg.cell(10, i, h)
    style_header_row(cfg, 10, 1, 6)

    ACC_START = 11
    for i, (name, typ, opening, limit, nw, notes) in enumerate(ACCOUNTS):
        r = ACC_START + i
        cfg.cell(r, 1, name).fill = yellow_fill
        cfg.cell(r, 1).border = thin
        cfg.cell(r, 2, typ).fill = yellow_fill
        cfg.cell(r, 2).border = thin
        money_cell(cfg.cell(r, 3), value=opening, editable=True)
        if limit is not None:
            money_cell(cfg.cell(r, 4), value=limit, editable=True)
        else:
            cfg.cell(r, 4).fill = yellow_fill
            cfg.cell(r, 4).border = thin
            cfg.cell(r, 4).number_format = inr
        cfg.cell(r, 5, "TRUE" if nw else "FALSE").fill = yellow_fill
        cfg.cell(r, 5).border = thin
        cfg.cell(r, 6, notes).fill = yellow_fill
        cfg.cell(r, 6).border = thin
    ACC_END = ACC_START + len(ACCOUNTS) - 1

    cfg["H8"] = "LEDGER TYPES"
    cfg["H8"].font = section_font
    cfg["H8"].fill = section_fill
    cfg["H9"] = "Type"
    cfg["H9"].font = header_font
    cfg["H9"].fill = header_fill
    for i, t in enumerate(TYPES):
        cfg.cell(10 + i, 8, t).fill = yellow_fill
        cfg.cell(10 + i, 8).border = thin
    TYPE_END = 10 + len(TYPES) - 1

    guides = [
        ("Income", "From=Employer → To=bank/cash. Budget=FALSE."),
        ("Expense", "From=account/card → To=Expense. Budget=TRUE for normal spends."),
        ("Transfer", "Between your accounts. Budget=FALSE."),
        ("Credit Card Payment", "From=savings → To=credit card. Budget=FALSE."),
        ("Refund", "From=Expense → To=account/card."),
        ("Investment", "From=savings → To=FD / Mutual Fund. Budget=FALSE."),
        ("Adjustment", "Reconciliation catch-up. Keeps ledger as source of truth."),
    ]
    cfg["I8"] = "TYPE GUIDE"
    cfg["I8"].font = section_font
    cfg["I9"] = "Type"
    cfg["J9"] = "From → To pattern"
    style_header_row(cfg, 9, 9, 10)
    for i, (t, g) in enumerate(guides):
        cfg.cell(10 + i, 9, t).border = thin
        cfg.cell(10 + i, 10, g).border = thin

    cfg["A24"] = "CATEGORIES"
    cfg["A24"].font = section_font
    cfg["A24"].fill = section_fill
    cfg.merge_cells("A24:C24")
    cfg["A25"] = (
        "Add new rows under the list (yellow). Group helps review. "
        "Typical Budget? = default for phone form / new Ledger rows — override per entry."
    )
    cfg["A25"].font = muted_font
    cfg.merge_cells("A25:C25")
    cfg["A26"] = "Category"
    cfg["B26"] = "Group"
    cfg["C26"] = "Typical Budget?"
    style_header_row(cfg, 26, 1, 3)
    CAT_START = 27
    for i, (cat, group, bud) in enumerate(CATEGORIES):
        r = CAT_START + i
        cfg.cell(r, 1, cat).fill = yellow_fill
        cfg.cell(r, 1).border = thin
        cfg.cell(r, 2, group).fill = yellow_fill
        cfg.cell(r, 2).border = thin
        cfg.cell(r, 3, "TRUE" if bud else "FALSE").fill = yellow_fill
        cfg.cell(r, 3).border = thin
    CAT_END = CAT_START + len(CATEGORIES) - 1
    # Extra blank yellow rows so you can add varieties without rebuilding
    EXTRA_CAT_ROWS = 15
    for r in range(CAT_END + 1, CAT_END + 1 + EXTRA_CAT_ROWS):
        for c in (1, 2, 3):
            cfg.cell(r, c).fill = yellow_fill
            cfg.cell(r, c).border = thin
    CAT_DV_END = CAT_END + EXTRA_CAT_ROWS

    rules_header = CAT_DV_END + 2
    cfg.cell(rules_header, 1, "RULES").font = section_font
    cfg.cell(rules_header, 1).fill = section_fill
    rules = [
        "1. Enter every money movement exactly once in the Ledger. Never edit calculated balances.",
        "2. From Account / To Account replace Payment Method + Credit Card.",
        "3. Amount is always positive. Direction is entirely From → To.",
        "4. Include in Budget = TRUE only for normal monthly spends under the budget cap.",
        "5. Rent, EMI, salary, transfers, investments, CC bill payments → usually Budget = FALSE.",
        "6. Opening balances on this sheet are the ONLY manual balance inputs (starting point).",
        "7. When calculated ≠ bank app, use Reconciliation → add missing Ledger row or Adjustment.",
        "8. Asset = Opening + To − From. Liability due = Opening + From − To.",
        "9. Prefer a specific category over Other — it makes Dashboard / review much easier.",
        "10. To add a category: type it in the next blank yellow row (Group + Typical Budget?). Phone form reads Configuration.",
    ]
    for i, rule in enumerate(rules):
        rr = rules_header + 1 + i
        cfg.cell(rr, 1, rule)
        cfg.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)

    for col, w in zip(
        list("ABCDEFHIJ"), [28, 16, 16, 18, 18, 36, 22, 22, 70]
    ):
        cfg.column_dimensions[col].width = w

    # ── Ledger ─────────────────────────────────────────────────────────
    led = wb.create_sheet("Ledger")
    headers = [
        "Date",
        "Time",
        "Day",
        "Month",
        "Year",
        "Type",
        "Amount (₹)",
        "From Account",
        "To Account",
        "Category",
        "Include in Budget",
        "Notes",
        "Source",  # manual (form) | ai (voice LLM)
    ]
    for i, h in enumerate(headers, 1):
        led.cell(1, i, h)
    style_header_row(led, 1, 1, 13)
    led.row_dimensions[1].height = 30
    led.freeze_panes = "A2"
    led.auto_filter.ref = "A1:M1000"

    samples = [
        (46240.0, 0.4409722222222222, "Expense", 35.0, "HDFC Savings", "Expense", "Eating outside", True, "Local food stall", "manual"),
        (46240.0, 0.45416666666666666, "Expense", 18.0, "HDFC Savings", "Expense", "Groceries - Physical", True, "Akshayakalpa milk", "manual"),
    ]
    for i, (dt_s, tm, typ, amt, fr, to, cat, bud, notes, src) in enumerate(samples):
        r = 2 + i
        led.cell(r, 1, dt_s).number_format = "dd/mm/yyyy"
        led.cell(r, 1).fill = yellow_fill
        led.cell(r, 2, tm).number_format = "HH:mm"
        led.cell(r, 2).fill = yellow_fill
        led.cell(r, 3, f'=IF(A{r}="","",TEXT(A{r},"dddd"))')
        led.cell(r, 4, f'=IF(A{r}="","",TEXT(A{r},"MMMM"))')
        led.cell(r, 5, f'=IF(A{r}="","",YEAR(A{r}))')
        led.cell(r, 6, typ).fill = yellow_fill
        money_cell(led.cell(r, 7), value=amt, editable=True)
        led.cell(r, 8, fr).fill = yellow_fill
        led.cell(r, 9, to).fill = yellow_fill
        led.cell(r, 10, cat).fill = yellow_fill
        led.cell(r, 11, bud).fill = yellow_fill
        led.cell(r, 12, notes).fill = yellow_fill
        led.cell(r, 13, src).fill = yellow_fill
        for c in range(1, 14):
            led.cell(r, c).border = thin

    for r in range(4, 202):
        led.cell(r, 3, f'=IF(A{r}="","",TEXT(A{r},"dddd"))')
        led.cell(r, 4, f'=IF(A{r}="","",TEXT(A{r},"MMMM"))')
        led.cell(r, 5, f'=IF(A{r}="","",YEAR(A{r}))')
        for c in [1, 2, 6, 7, 8, 9, 10, 11, 12, 13]:
            led.cell(r, c).fill = yellow_fill
            led.cell(r, c).border = thin
        led.cell(r, 7).number_format = inr
        led.cell(r, 1).number_format = "dd/mm/yyyy"
        led.cell(r, 2).number_format = "HH:mm"

    dv_type = DataValidation(
        type="list", formula1=f"Configuration!$H$10:$H${TYPE_END}", allow_blank=True
    )
    dv_acc = DataValidation(
        type="list", formula1=f"Configuration!$A${ACC_START}:$A${ACC_END}", allow_blank=True
    )
    dv_cat = DataValidation(
        type="list",
        formula1=f"Configuration!$A${CAT_START}:$A${CAT_DV_END}",
        allow_blank=True,
    )
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_source = DataValidation(type="list", formula1='"manual,ai"', allow_blank=True)
    for dv in (dv_type, dv_acc, dv_cat, dv_bool, dv_source):
        led.add_data_validation(dv)
    dv_type.add("F2:F1000")
    dv_acc.add("H2:H1000")
    dv_acc.add("I2:I1000")
    dv_cat.add("J2:J1000")
    dv_bool.add("K2:K1000")
    dv_source.add("M2:M1000")

    for col, w in zip(
        list("ABCDEFGHIJKLM"), [12, 8, 12, 12, 8, 20, 12, 18, 18, 22, 16, 28, 10]
    ):
        led.column_dimensions[col].width = w

    # ── Monthly Budget ─────────────────────────────────────────────────
    mb = wb.create_sheet("Monthly Budget")
    mb["A1"] = "MONTHLY BUDGET"
    mb["A1"].font = title_font
    mb["A2"] = (
        "Only the Budget column is manual (yellow). All other figures come from the Ledger."
    )
    mb["A2"].font = muted_font
    mb.merge_cells("A2:L2")

    mb["A4"] = "THIS MONTH"
    mb["A4"].font = section_font
    mb["A4"].fill = section_fill
    mb["A5"] = "Month"
    mb["B5"] = '=TEXT(TODAY(),"MMMM YYYY")'
    mb["B5"].font = Font(bold=True, size=14, color="1F4E79")

    L = "Ledger"
    mb["A6"] = "Budget"
    money_cell(
        mb["B6"],
        formula='=IFERROR(INDEX($B$20:$B$67,MATCH(DATE(YEAR(TODAY()),MONTH(TODAY()),1),$A$20:$A$67,0)),Configuration!B6)',
    )
    rows_snap = [
        (
            "Income / Salary",
            f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$F:$F,"Income")',
        ),
        (
            "Budget Expenses",
            f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$F:$F,"Expense",{L}!$K:$K,TRUE)'
            f'-SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$F:$F,"Refund",{L}!$K:$K,TRUE)',
        ),
        ("Remaining Budget", "=B6-B8"),
        (
            "Non-Budget Expenses",
            f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$F:$F,"Expense",{L}!$K:$K,FALSE)',
        ),
        (
            "Investments",
            f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$F:$F,"Investment")',
        ),
        (
            "EMIs",
            f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$J:$J,"EMIs")',
        ),
        (
            "Rent",
            f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$J:$J,"Rent")',
        ),
        (
            "CC Payments",
            f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(TODAY(),"MMMM"),{L}!$E:$E,YEAR(TODAY()),{L}!$F:$F,"Credit Card Payment")',
        ),
        ("Est. Savings (Income − outflows)", "=B7-B8-B10-B11"),
    ]
    for i, (label, formula) in enumerate(rows_snap):
        r = 7 + i
        mb.cell(r, 1, label).border = thin
        money_cell(mb.cell(r, 2), formula=formula)
    mb["B9"].fill = good_fill
    mb["A16"] = "Budget used %"
    mb["B16"] = "=IF(B6=0,0,B8/B6)"
    mb["B16"].number_format = pct
    mb["B16"].fill = calc_fill
    mb["B16"].border = thin

    mb["A18"] = "MONTH-BY-MONTH — only Budget (col B) is editable"
    mb["A18"].font = section_font
    mb["A18"].fill = section_fill
    mb.merge_cells("A18:L18")
    for i, h in enumerate(
        [
            "Month Start",
            "Budget",
            "Income",
            "Budget Expenses",
            "Remaining",
            "Non-Budget Exp",
            "Investments",
            "EMIs",
            "Rent",
            "CC Payments",
            "Est. Savings",
            "Label",
        ],
        1,
    ):
        mb.cell(19, i, h)
    style_header_row(mb, 19, 1, 12)

    start = ddate(2026, 1, 1)
    for i in range(48):
        r = 20 + i
        y = start.year + (start.month - 1 + i) // 12
        m = (start.month - 1 + i) % 12 + 1
        mb.cell(r, 1, ddate(y, m, 1)).number_format = "mmm yyyy"
        mb.cell(r, 1).border = thin
        money_cell(mb.cell(r, 2), value=10000, editable=True)
        money_cell(
            mb.cell(r, 3),
            formula=f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$F:$F,"Income")',
        )
        money_cell(
            mb.cell(r, 4),
            formula=f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$F:$F,"Expense",{L}!$K:$K,TRUE)'
            f'-SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$F:$F,"Refund",{L}!$K:$K,TRUE)',
        )
        money_cell(mb.cell(r, 5), formula=f"=B{r}-D{r}")
        money_cell(
            mb.cell(r, 6),
            formula=f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$F:$F,"Expense",{L}!$K:$K,FALSE)',
        )
        money_cell(
            mb.cell(r, 7),
            formula=f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$F:$F,"Investment")',
        )
        money_cell(
            mb.cell(r, 8),
            formula=f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$J:$J,"EMIs")',
        )
        money_cell(
            mb.cell(r, 9),
            formula=f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$J:$J,"Rent")',
        )
        money_cell(
            mb.cell(r, 10),
            formula=f'=SUMIFS({L}!$G:$G,{L}!$D:$D,TEXT(A{r},"MMMM"),{L}!$E:$E,YEAR(A{r}),{L}!$F:$F,"Credit Card Payment")',
        )
        money_cell(mb.cell(r, 11), formula=f"=C{r}-D{r}-F{r}-G{r}")
        mb.cell(r, 12, f'=TEXT(A{r},"MMMM YYYY")').border = thin

    for col, w in zip(list("ABCDEFGHIJKL"), [12, 12, 12, 15, 12, 14, 12, 10, 10, 12, 12, 14]):
        mb.column_dimensions[col].width = w

    # ── Reconciliation ─────────────────────────────────────────────────
    rec = wb.create_sheet("Reconciliation")
    rec["A1"] = "ACCOUNT RECONCILIATION"
    rec["A1"].font = title_font
    rec["A2"] = (
        "Calculated balance always comes from Opening + Ledger. Enter Actual only when you check "
        "your bank/app. If Difference ≠ 0: find the missing txn, or add Type=Adjustment. "
        "Never overwrite Calculated."
    )
    rec["A2"].font = muted_font
    rec.merge_cells("A2:G2")
    for i, h in enumerate(
        ["Account", "Type", "Calculated", "Actual", "Difference", "Last Reconciled", "Notes"],
        1,
    ):
        rec.cell(4, i, h)
    style_header_row(rec, 4, 1, 7)

    for i, _acc in enumerate(ACCOUNTS):
        r = 5 + i
        cfg_r = ACC_START + i
        rec.cell(r, 1, f"=Configuration!A{cfg_r}").border = thin
        rec.cell(r, 2, f"=Configuration!B{cfg_r}").border = thin
        calc = (
            f'=IF(B{r}="Liability",'
            f"Configuration!C{cfg_r}+SUMIF(Ledger!$H:$H,A{r},Ledger!$G:$G)-SUMIF(Ledger!$I:$I,A{r},Ledger!$G:$G),"
            f"Configuration!C{cfg_r}+SUMIF(Ledger!$I:$I,A{r},Ledger!$G:$G)-SUMIF(Ledger!$H:$H,A{r},Ledger!$G:$G))"
        )
        money_cell(rec.cell(r, 3), formula=calc)
        money_cell(rec.cell(r, 4), value=None, editable=True)
        money_cell(rec.cell(r, 5), formula=f'=IF(D{r}="","",D{r}-C{r})')
        rec.cell(r, 6).fill = yellow_fill
        rec.cell(r, 6).border = thin
        rec.cell(r, 6).number_format = "dd/mm/yyyy"
        rec.cell(r, 7).fill = yellow_fill
        rec.cell(r, 7).border = thin

    rec["A18"] = "HOW TO FIX A DIFFERENCE"
    rec["A18"].font = section_font
    for i, line in enumerate(
        [
            "1. Investigate bank statement / UPI history for the gap.",
            "2. If you find the missing txn → add it to the Ledger (do not edit Calculated).",
            "3. If you cannot find it → Ledger: Type=Adjustment, Category=Reconciliation, Notes=explain.",
            "4. After Ledger update, Calculated should match Actual. Set Last Reconciled = today.",
            '5. Never type over a balance to "make it right" — that destroys history.',
        ]
    ):
        rec.cell(19 + i, 1, line)

    for col, w in zip(list("ABCDEFG"), [20, 12, 14, 14, 12, 16, 40]):
        rec.column_dimensions[col].width = w

    # ── Planned Expenses (planning only — never Ledger) ──────────────
    pe = wb.create_sheet("Planned Expenses")
    populate_planned_expenses(pe, seed=True)

    # ── Dashboards ─────────────────────────────────────────────────────
    simple = wb.create_sheet("Simple Dashboard")
    populate_simple_dashboard(simple)
    detailed = wb.create_sheet("Detailed Dashboard")
    populate_detailed_dashboard(detailed, n_accounts=len(ACCOUNTS))

    wb._sheets = [simple, detailed, led, mb, rec, pe, cfg]
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print("Sheets:", wb.sheetnames)
    return OUT


def patch_live_categories(wb) -> tuple[int, int]:
    """Rewrite Configuration categories (with Group) + blank rows; move RULES below.
    Returns (cat_start, cat_dv_end) for Ledger validation.
    """
    cfg = wb["Configuration"]
    # Drop merges in the categories/rules zone first (MergedCell is read-only)
    for mr in list(cfg.merged_cells.ranges):
        if mr.min_row >= 24:
            cfg.unmerge_cells(str(mr))
    # Clear old categories + rules block
    for r in range(24, 100):
        for c in range(1, 7):
            cell = cfg.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()

    cfg["A24"] = "CATEGORIES"
    cfg["A24"].font = section_font
    cfg["A24"].fill = section_fill
    cfg.merge_cells("A24:C24")
    cfg["A25"] = (
        "Add new rows under the list (yellow). Group helps review. "
        "Typical Budget? = default hint — override per Ledger entry."
    )
    cfg["A25"].font = muted_font
    cfg.merge_cells("A25:C25")
    cfg["A26"] = "Category"
    cfg["B26"] = "Group"
    cfg["C26"] = "Typical Budget?"
    style_header_row(cfg, 26, 1, 3)
    CAT_START = 27
    for i, (cat, group, bud) in enumerate(CATEGORIES):
        r = CAT_START + i
        cfg.cell(r, 1, cat).fill = yellow_fill
        cfg.cell(r, 1).border = thin
        cfg.cell(r, 2, group).fill = yellow_fill
        cfg.cell(r, 2).border = thin
        cfg.cell(r, 3, "TRUE" if bud else "FALSE").fill = yellow_fill
        cfg.cell(r, 3).border = thin
    CAT_END = CAT_START + len(CATEGORIES) - 1
    EXTRA_CAT_ROWS = 15
    for r in range(CAT_END + 1, CAT_END + 1 + EXTRA_CAT_ROWS):
        for c in (1, 2, 3):
            cfg.cell(r, c).fill = yellow_fill
            cfg.cell(r, c).border = thin
    CAT_DV_END = CAT_END + EXTRA_CAT_ROWS

    rules_header = CAT_DV_END + 2
    cfg.cell(rules_header, 1, "RULES").font = section_font
    cfg.cell(rules_header, 1).fill = section_fill
    rules = [
        "1. Enter every money movement exactly once in the Ledger. Never edit calculated balances.",
        "2. From Account / To Account replace Payment Method + Credit Card.",
        "3. Amount is always positive. Direction is entirely From → To.",
        "4. Include in Budget = TRUE only for normal monthly spends under the budget cap.",
        "5. Rent, EMI, salary, transfers, investments, CC bill payments → usually Budget = FALSE.",
        "6. Opening balances on this sheet are the ONLY manual balance inputs (starting point).",
        "7. When calculated ≠ bank app, use Reconciliation → add missing Ledger row or Adjustment.",
        "8. Asset = Opening + To − From. Liability due = Opening + From − To.",
        "9. Prefer a specific category over Other — Dashboard review becomes much easier.",
        "10. To add a category: fill the next blank yellow row (name + Group + Typical Budget?).",
    ]
    for i, rule in enumerate(rules):
        rr = rules_header + 1 + i
        cfg.cell(rr, 1, rule)
        cfg.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)

    cfg.column_dimensions["A"].width = 28
    cfg.column_dimensions["B"].width = 16
    cfg.column_dimensions["C"].width = 16
    return CAT_START, CAT_DV_END


def patch_live_ledger_category_validation(wb, cat_start: int, cat_dv_end: int) -> None:
    led = wb["Ledger"]
    # Replace category list validation
    keep = []
    for dv in list(led.data_validations.dataValidation):
        sq = str(dv.sqref)
        # Drop old category validators on column J
        if "J" in sq and dv.type == "list" and dv.formula1 and "Configuration!$A$" in str(dv.formula1):
            continue
        keep.append(dv)
    led.data_validations.dataValidation = keep
    dv_cat = DataValidation(
        type="list",
        formula1=f"Configuration!$A${cat_start}:$A${cat_dv_end}",
        allow_blank=True,
    )
    led.add_data_validation(dv_cat)
    dv_cat.add("J2:J1000")


def _n_accounts_from_wb(wb) -> int:
    n_accounts = len(ACCOUNTS)
    if "Configuration" in wb.sheetnames:
        cfg = wb["Configuration"]
        count = 0
        for r in range(11, 40):
            if cfg.cell(r, 1).value:
                count += 1
            else:
                break
        if count:
            n_accounts = count
    return n_accounts


def _recreate_sheet(wb, name: str, preferred_index: int = 0):
    """Delete sheet if present and recreate at preferred_index (or append)."""
    if name in wb.sheetnames:
        del wb[name]
    # Clamp index to current sheet count
    idx = min(preferred_index, len(wb.sheetnames))
    return wb.create_sheet(name, idx)


def patch_live_monthly_salary(wb) -> None:
    """Ensure Configuration has Monthly Salary at B7 (does not shift account rows)."""
    cfg = wb["Configuration"]
    # Row 7 sits between Budget default (6) and ACCOUNTS header (8) — safe insert point
    label = str(cfg["A7"].value or "").strip().lower()
    if "salary" not in label:
        cfg["A7"] = "Monthly Salary (₹)"
        cfg["A7"].border = thin
        # Preserve any existing number the user may have typed into B7 already
        existing = cfg["B7"].value
        if existing is None or (isinstance(existing, str) and not str(existing).strip()):
            money_cell(cfg["B7"], value=0, editable=True)
        else:
            money_cell(cfg["B7"], value=existing, editable=True)
        cfg["C7"] = "Take-home used for next-month free-to-allocate estimate on dashboards"
        cfg["C7"].font = muted_font
    elif cfg["B7"].value is None:
        money_cell(cfg["B7"], value=0, editable=True)


def patch_live_dashboard(path: Path | None = None) -> Path:
    """Rebuild Simple + Detailed dashboards; refresh categories + salary config."""
    path = path or LIVE
    if not path.exists():
        raise SystemExit(f"Live workbook not found: {path}")
    wb = load_workbook(path)

    patch_live_monthly_salary(wb)
    cat_start, cat_dv_end = patch_live_categories(wb)
    patch_live_ledger_category_validation(wb, cat_start, cat_dv_end)

    # Drop legacy single "Dashboard" if still present
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]

    # Ensure Planned Expenses exists; do not wipe existing planning data on rebuild
    if "Planned Expenses" not in wb.sheetnames:
        pe = wb.create_sheet("Planned Expenses")
        populate_planned_expenses(pe, seed=True)
        print("Created Planned Expenses sheet (seeded sample rows)")

    n_accounts = _n_accounts_from_wb(wb)
    simple = _recreate_sheet(wb, "Simple Dashboard", 0)
    populate_simple_dashboard(simple)
    detailed = _recreate_sheet(wb, "Detailed Dashboard", 1)
    populate_detailed_dashboard(detailed, n_accounts=n_accounts)

    # Preferred order: Simple first, then Detailed, then the rest
    preferred = ["Simple Dashboard", "Detailed Dashboard"]
    rest = [s for s in wb.sheetnames if s not in preferred]
    order = preferred + rest
    wb._sheets = [wb[s] for s in order if s in wb.sheetnames]
    wb.save(path)
    print(f"Patched Simple + Detailed dashboards + categories on {path}")
    print(f"Categories: {len(CATEGORIES)} (+15 blank slots), validation A{cat_start}:A{cat_dv_end}")
    print(f"Charts on Detailed Dashboard: {len(detailed._charts)}")
    print("Simple free-to-allocate formula:", simple["B20"].value)
    print("Simple next-month free formula:", simple["B28"].value)
    print("Config monthly salary (B7):", wb["Configuration"]["B7"].value)
    return path


def ensure_planned_expenses_live(path: Path | None = None, *, seed: bool = True) -> Path:
    """
    Surgical live update:
      - create/refresh Planned Expenses sheet (seed only if new or empty)
      - append Planned Expenses summary block on Simple + Detailed dashboards
        without recreating those sheets.
    Does not touch Ledger, Monthly Budget, Reconciliation, or Configuration.
    """
    path = path or LIVE
    if not path.exists():
        raise SystemExit(f"Live workbook not found: {path}")
    wb = load_workbook(path)

    created = False
    if "Planned Expenses" in wb.sheetnames:
        pe = wb["Planned Expenses"]
        # Refresh layout only if sheet looks empty (no title)
        if not pe["A1"].value:
            populate_planned_expenses(pe, seed=seed)
            created = True
        # else keep user data as-is
    else:
        pe = wb.create_sheet("Planned Expenses")
        populate_planned_expenses(pe, seed=seed)
        created = True

    # Simple Dashboard — find free row after existing content
    if "Simple Dashboard" in wb.sheetnames:
        simple = wb["Simple Dashboard"]
        # Avoid duplicating if already linked
        already = False
        for r in range(1, min(simple.max_row or 1, 80) + 1):
            v = simple.cell(r, 1).value
            if v and "PLANNED EXPENSES" in str(v).upper():
                already = True
                break
        if not already:
            last = 1
            for r in range(1, 80):
                if any(simple.cell(r, c).value is not None for c in range(1, 5)):
                    last = r
            start = last + 2
            _add_planned_summary_block(simple, start, cols="D")
            print(f"Simple Dashboard: Planned Expenses block at row {start}")
        else:
            print("Simple Dashboard: Planned Expenses block already present")

    # Detailed Dashboard — place below existing content / charts note
    if "Detailed Dashboard" in wb.sheetnames:
        detailed = wb["Detailed Dashboard"]
        already = False
        for r in range(1, min(detailed.max_row or 1, 100) + 1):
            v = detailed.cell(r, 1).value
            if v and "PLANNED EXPENSES" in str(v).upper():
                already = True
                break
        if not already:
            last = 1
            for r in range(1, 100):
                if any(detailed.cell(r, c).value is not None for c in range(1, 7)):
                    last = r
            start = max(last + 2, 45)
            _add_planned_summary_block(detailed, start, cols="F")
            print(f"Detailed Dashboard: Planned Expenses block at row {start}")
        else:
            print("Detailed Dashboard: Planned Expenses block already present")

    # Prefer Planned Expenses after Reconciliation in tab order if newly created
    preferred = [
        "Simple Dashboard",
        "Detailed Dashboard",
        "Ledger",
        "Monthly Budget",
        "Reconciliation",
        "Planned Expenses",
        "Configuration",
    ]
    rest = [s for s in wb.sheetnames if s not in preferred]
    order = [s for s in preferred if s in wb.sheetnames] + rest
    wb._sheets = [wb[s] for s in order]

    wb.save(path)
    print(f"Planned Expenses ensured on {path} (created/refreshed={created})")
    print("Sheets:", wb.sheetnames)
    return path


def _atomic_save_live(wb, path: Path) -> None:
    import os

    tmp = path.with_suffix(".xlsx.writing")
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    except Exception:
        try:
            if tmp.is_file():
                tmp.unlink()
        except OSError:
            pass
        raise


def _last_used_row(ws, max_scan: int = 120, max_col: int = 6) -> int:
    last = 1
    for r in range(1, max_scan + 1):
        if any(ws.cell(r, c).value is not None for c in range(1, max_col + 1)):
            last = r
    return last


def ensure_commitment_forecast_live(path: Path | None = None) -> Path:
    """
    Surgical live update (does NOT recreate dashboards):
      - Kind column on Planned Expenses + classify existing rows
      - NEXT 6 MONTHS cash-due table (L4:P13)
      - Monthly Fixed Cost (B5) = this month cash due (P6)
      - Append 6-month table on Simple + Detailed if missing
      - Refresh MacBook SmartEMI notes from the HDFC letter
    Does not touch Ledger, Monthly Budget, Reconciliation, or Configuration.
    """
    path = path or LIVE
    if not path.exists():
        raise SystemExit(f"Live workbook not found: {path}")
    wb = load_workbook(path)
    if "Planned Expenses" not in wb.sheetnames:
        raise SystemExit("No Planned Expenses sheet — run --planned-expenses first")

    pe = wb["Planned Expenses"]
    kind_col = _ensure_kind_column(pe)
    _write_planned_forecast_table(pe, kind_col)

    # This month cash due (date-aware), not yearly÷12 of every Active row
    pe["B5"] = "=P6"
    pe["B5"].font = big_num_font
    pe["B5"].fill = soft_fill
    pe["B5"].number_format = inr
    pe["A11"] = (
        "Monthly Fixed Cost = this month's recurring cash due (P6). "
        "Kind splits Loan / EMI vs Lifestyle vs Investment. "
        "Start/End dates include a month if they overlap it. "
        "Yearly amounts count in the due month only."
    )

    # MacBook SmartEMI — keep amount ₹38,200; document letter details in Notes
    notes_col = None
    for c in range(1, 16):
        v = pe.cell(14, c).value
        if v and str(v).strip().lower() == "notes":
            notes_col = c
            break
    for r in range(15, 65):
        name = str(pe.cell(r, 1).value or "")
        if "smartemi" not in name.lower() and "macbook" not in name.lower():
            continue
        pe.cell(r, kind_col, PE_KIND_LOAN)
        if notes_col:
            pe.cell(r, notes_col).value = (
                "HDFC SmartEMI 144274447 · principal ₹2,19,500 · 6 mo @ 1.25% p.m. "
                "· billed EMI ₹38,200 (first due 12-Sep-2026 is ₹40,121; last ₹38,203) "
                "· GST 18% on interest extra · HDFC CC"
            )
        break

    def _refresh_fixed_cost_caption(dash) -> None:
        for r in range(1, min(dash.max_row or 1, 80) + 1):
            if str(dash.cell(r, 1).value or "").strip().lower() != "monthly fixed cost":
                continue
            note = str(dash.cell(r, 3).value or "")
            if "monthly equivalent" in note.lower() or not note:
                dash.cell(r, 3).value = "this month cash due (Start/End + Kind)"
                dash.cell(r, 3).font = muted_font

    if "Simple Dashboard" in wb.sheetnames:
        simple = wb["Simple Dashboard"]
        _refresh_fixed_cost_caption(simple)
        if not _dashboard_has_block(simple, "NEXT 6 MONTHS"):
            start = _last_used_row(simple, 80, 5) + 2
            _add_commitment_forecast_block(simple, start, cols="E")
            print(f"Simple Dashboard: NEXT 6 MONTHS block at row {start}")
        else:
            print("Simple Dashboard: NEXT 6 MONTHS block already present — not rewritten")

    if "Detailed Dashboard" in wb.sheetnames:
        detailed = wb["Detailed Dashboard"]
        _refresh_fixed_cost_caption(detailed)
        if not _dashboard_has_block(detailed, "NEXT 6 MONTHS"):
            # Stay below the category chart (~rows 55–73)
            start = max(_last_used_row(detailed, 100, 6) + 2, 75)
            _add_commitment_forecast_block(detailed, start, cols="E")
            print(f"Detailed Dashboard: NEXT 6 MONTHS block at row {start}")
        else:
            print("Detailed Dashboard: NEXT 6 MONTHS block already present — not rewritten")

    _atomic_save_live(wb, path)
    print(f"Commitment forecast ensured on {path}")
    print(f"Kind column: {get_column_letter(kind_col)}")
    return path


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] in {"--patch-live", "--live", "patch"}:
        patch_live_dashboard()
    elif len(sys.argv) > 1 and sys.argv[1] in {
        "--planned-expenses",
        "--ensure-planned",
        "planned",
    }:
        ensure_planned_expenses_live()
    elif len(sys.argv) > 1 and sys.argv[1] in {
        "--commitment-forecast",
        "--forecast",
        "forecast",
    }:
        ensure_commitment_forecast_live()
    else:
        build()
